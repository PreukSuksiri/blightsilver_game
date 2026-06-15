#!/usr/bin/env python3
"""Clone card_data.xlsx with grammar-only fixes. No style, terminology, or logic changes."""
from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "context" / "card_data.xlsx"
TS = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
OUT = ROOT / "context" / f"card_data_grammar_{TS}.xlsx"


def fix_grammar(text: str | None) -> tuple[str, list[str]]:
    """Return grammar-corrected text and a list of fix labels applied."""
    if text is None:
        return "", []
    t = str(text).strip()
    if not t or t.lower() == "none":
        return t, []

    original = t
    notes: list[str] = []

    def sub(pat: str, rep: str, note: str, *, flags: int = re.I) -> None:
        nonlocal t
        new = re.sub(pat, rep, t, flags=flags)
        if new != t:
            if note not in notes:
                notes.append(note)
            t = new

    # Broken typos
    sub(r"\b on the this player", " on this player", "broken this player typo")
    sub(r"\bthe the\b", "the", "duplicate the")
    sub(r"\ba a\b", "a", "duplicate a")
    sub(r"\bis is\b", "is", "duplicate is")
    sub(r":\s*,", ":", "colon comma typo")
    sub(r":{2,}", ":", "double colon typo")

    # Subject–verb agreement
    sub(r"\bit attack\b", "it attacks", "subject–verb agreement")
    sub(r"\bit gain\b", "it gains", "subject–verb agreement")
    sub(r"\bit lose\b", "it loses", "subject–verb agreement")
    sub(r"\bit perform\b", "it performs", "subject–verb agreement")
    sub(r"\bit get\b", "it gets", "subject–verb agreement")
    sub(r"\bit have\b", "it has", "subject–verb agreement")
    sub(r"\bowner lose\b", "owner loses", "subject–verb agreement")
    sub(r"\bowner gain\b", "owner gains", "subject–verb agreement")
    sub(r"\bthis card gain\b", "this card gains", "subject–verb agreement")
    sub(r"\bfoe lose\b", "foe loses", "subject–verb agreement")
    sub(r"\bfoe choose\b", "foe chooses", "subject–verb agreement")
    sub(r"\bfoe must chooses\b", "foe must choose", "subject–verb agreement")
    sub(r"\bfoe don't\b", "foe doesn't", "subject–verb agreement")
    sub(r"\bfoe pay no cost\b", "foe pays no cost", "subject–verb agreement")
    sub(r"\bFoe pay no cost\b", "Foe pays no cost", "subject–verb agreement")
    sub(r"\battacker have\b", "attacker has", "subject–verb agreement")
    sub(r"\bThe attacker choose\b", "The attacker chooses", "subject–verb agreement")
    sub(r"(?<![Tt]he )\battacker choose\b", "The attacker chooses", "subject–verb agreement")
    sub(r"\bAttacker choose\b", "The attacker chooses", "subject–verb agreement")
    sub(r"\bAttacker end\b", "Attacker ends", "subject–verb agreement")
    sub(r"\btrapper have\b", "trapper has", "subject–verb agreement")
    sub(r"\bThe trapper choose\b", "The trapper chooses", "subject–verb agreement")
    sub(r"(?<![Tt]he )\btrapper choose\b", "The trapper chooses", "subject–verb agreement")
    sub(r"\bTrapper choose\b", "The trapper chooses", "subject–verb agreement")
    sub(r"\bthe owner gain\b", "the owner gains", "subject–verb agreement")
    sub(r"\bthe owner use\b", "the owner uses", "subject–verb agreement")
    sub(r"\bthe owner have\b", "the owner has", "subject–verb agreement")
    sub(r"\bthe owner can chooses\b", "the owner can choose", "subject–verb agreement")
    sub(r"\bThis player choose\b", "This player chooses", "subject–verb agreement")
    sub(r"\bThis player select\b", "This player selects", "subject–verb agreement")
    sub(r"\bboth player\b", "both players", "subject–verb agreement")
    sub(r"\bThis bonus do not\b", "This bonus does not", "subject–verb agreement")
    sub(r"\bThose monster\b", "That monster", "subject–verb agreement")
    sub(r"\bthose monster\b", "that monster", "subject–verb agreement")

    # Verb form / tense
    sub(r"\bis destroy\b", "is destroyed", "verb form")
    sub(r"\bwill be destroy\b", "would be destroyed", "verb form")
    sub(r"\bAfter it is destroy\b", "After it is destroyed", "verb form")
    sub(r"\bAfter successfully attack\b", "After successfully attacking", "verb form")
    sub(r"\bAfter performed an attack\b", "After performing an attack", "verb form")

    # Missing articles
    sub(r"\bAfter successful attack\b", "After a successful attack", "missing article")
    sub(r"\bflip coin\b", "flip a coin", "missing article")
    sub(r"\bFlip coin\b", "Flip a coin", "missing article")
    sub(r"\bif a exposed\b", "if an exposed", "missing article")
    sub(r"\bIf a exposed\b", "If an exposed", "missing article")

    # Possessives
    sub(r"\bat end of owner turn\b", "at end of owner's turn", "possessive")
    sub(r"\bat end of foe turn\b", "at end of foe's turn", "possessive")
    sub(r"\beach owner turn\b", "each owner's turn", "possessive")
    sub(r"\buntil foe turn ends\b", "until foe's turn ends", "possessive")
    sub(r"\bUntil foe turn ends\b", "Until foe's turn ends", "possessive")
    sub(r"\bowner turn start\b", "owner's turn start", "possessive")
    sub(r"\bfoe turn start\b", "foe's turn start", "possessive")
    sub(r"\bowner turn end\b", "owner's turn end", "possessive")
    sub(r"\bfoe turn end\b", "foe's turn end", "possessive")

    # Missing subject / object
    sub(r"\bif survived,\b", "if this card survived,", "missing subject")
    sub(r"\bIf survived,\b", "If this card survived,", "missing subject")
    sub(r"\bthe attacker attack is negated\b", "the attacker's attack is negated", "missing possessive")
    sub(r"\battacker attack is negated\b", "the attacker's attack is negated", "missing possessive")

    # Plural agreement
    sub(r"\bcells on its side is revealed\b", "cells on its side are revealed", "plural agreement")
    sub(r"\bcells on its side is\b", "cells on its side are", "plural agreement")
    sub(r"\bor more of cells on its side is revealed\b", "or more cells on its side are revealed", "plural agreement")
    sub(r"\bIf at least 1 (.+?) on field\b", r"If at least 1 \1 is on the field", "missing verb")
    sub(r"\bhave (\d+) or less card on the field\b", r"has \1 or less cards on the field", "subject–verb agreement")

    # Whitespace and sentence start
    t = re.sub(r"  +", " ", t).strip()
    if t and t[0].islower():
        t = t[0].upper() + t[1:]
        if "sentence start" not in notes and t != original:
            notes.append("sentence start")

    if t == original and not notes:
        return original, []
    if not notes:
        notes.append("grammar")
    return t, notes


def process_simple_sheet(ws, ability_col: int, old_col: int, comment_col: int) -> None:
    ws.cell(row=2, column=old_col, value="Old Ability")
    ws.cell(row=2, column=comment_col, value="Comment")
    for row in range(3, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name:
            continue
        old = ws.cell(row=row, column=ability_col).value
        if old is None or str(old).strip() == "":
            continue
        old_s = str(old).strip()
        if old_s.lower() == "none":
            continue
        new, notes = fix_grammar(old_s)
        if new != old_s:
            ws.cell(row=row, column=old_col, value=old_s)
            ws.cell(row=row, column=ability_col, value=new)
            ws.cell(row=row, column=comment_col, value="; ".join(notes))


def process_union_sheet(
    ws,
    partial_col: int,
    full_col: int,
    old_partial_col: int,
    old_full_col: int,
    comment_col: int,
) -> None:
    ws.cell(row=2, column=old_partial_col, value="Old Partial Ability")
    ws.cell(row=2, column=old_full_col, value="Old Full Ability")
    ws.cell(row=2, column=comment_col, value="Comment")
    for row in range(3, ws.max_row + 1):
        if not ws.cell(row=row, column=1).value:
            continue
        comments: list[str] = []

        partial_old = ws.cell(row=row, column=partial_col).value
        if partial_old and str(partial_old).strip().lower() not in ("none", ""):
            partial_s = str(partial_old).strip()
            partial_new, notes = fix_grammar(partial_s)
            if partial_new != partial_s:
                ws.cell(row=row, column=old_partial_col, value=partial_s)
                ws.cell(row=row, column=partial_col, value=partial_new)
                comments.append("Partial: " + "; ".join(notes))

        full_old = ws.cell(row=row, column=full_col).value
        if full_old and str(full_old).strip().lower() not in ("none", ""):
            full_s = str(full_old).strip()
            full_new, notes = fix_grammar(full_s)
            if full_new != full_s:
                ws.cell(row=row, column=old_full_col, value=full_s)
                ws.cell(row=row, column=full_col, value=full_new)
                comments.append("Full: " + "; ".join(notes))

        if comments:
            ws.cell(row=row, column=comment_col, value=" | ".join(comments))


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"Missing source: {SRC}")
    shutil.copy2(SRC, OUT)
    wb = openpyxl.load_workbook(OUT)

    for sheet in ("Unit", "Trap", "Tech"):
        ws = wb[sheet]
        ws.insert_cols(9, amount=2)  # Old Ability, Comment after Ability (col 8)
        process_simple_sheet(ws, ability_col=8, old_col=9, comment_col=10)

    ws_u = wb["Union"]
    ws_u.insert_cols(10, amount=3)
    process_union_sheet(
        ws_u,
        partial_col=8,
        full_col=9,
        old_partial_col=10,
        old_full_col=11,
        comment_col=12,
    )

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
