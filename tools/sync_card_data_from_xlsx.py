#!/usr/bin/env python3
"""Sync CardDatabase.gd, UnionDatabase.gd, and demo_flags.json from context/card_data.xlsx."""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "context" / "card_data.xlsx"
CARD_DB = ROOT / "autoload" / "CardDatabase.gd"
UNION_DB = ROOT / "autoload" / "UnionDatabase.gd"
DEMO_FLAGS = ROOT / "data" / "demo_flags.json"
CUSTOM_PACKS = ROOT / "shop" / "custom_packs.json"

NAME_ALIASES = {
    "Armored Money": "Armored Monkey",
}

# xlsx formula / legacy references → current in-game card name
CARD_RENAMES = {
    "Alluring Spellcaster": "Alluring Witch",
    "Century Fortuneteller": "Uncanny Fortuneteller",
    "Deep Tribe Axe Thrower": "Deep Tribe Axeman",
    "Freya the Rift Walker": "Freya the Rift Hunter",
    "Gremlin Worker": "Hardworking Gremlin",
    "Library Critters": "Library Critter",
    "Lindsy the Brave Princess": "Chihako the Brave Princess",
    "Magical Smith": "Magical Craftsman",
    "Mystic Dagger Dancer": "Mystic Dancer",
    "Ore Transporter": "Ore Eater",
    "Outpost Magician": "Outpost Mage",
    "Psy Warrior": "Mystic Warrior",
    "Rotten Dog": "Skeleton Dog",
    "Spell Exile": "Spellcaster Exile",
    "Chronoteleporter": "Chronomancer",
}

PACK_ID_BY_SHORT_NAME = {
    "Blightsilver": "booster_blightsilver",
    "Unseen Presence": "booster_unseen_presence",
    "Wind of Dominance": "booster_wind_of_dominance",
}

SHEET_CARD_TYPE = {
    "Unit": "character",
    "Trap": "trap",
    "Tech": "tech",
}

WEIGHT_ALIASES = {
    "Night Whisperer": "Night Whisperor",
    "Moon Tribe Twin Blader": "Moon Tribe Twin Blades",
}

RARITY_DEFAULT_WEIGHT = {
    1: 100,
    2: 50,
    3: 20,
    4: 5,
    5: 1,
}

AFFINITY = {
    "Divine": "DIVINE",
    "Chaos": "CHAOS",
    "Nature": "NATURE",
    "Arcane": "ARCANE",
    "Cosmic": "COSMIC",
    "Bio": "BIO",
    "Anima": "ANIMA",
    "Machine": "MACHINE",
}

AFFINITY_WORDS = {k.lower(): v for k, v in AFFINITY.items()}
AFFINITY_WORDS["cosmo"] = "COSMIC"

# xlsx rows with missing/typo affinity (card name → sheet Affinity value)
UNIT_AFFINITY_OVERRIDES: dict[str, str] = {
    "Deep Tribe Witchdoctor": "Nature",
}


def resolve_affinity(raw, *, card_name: str = "") -> str | None:
    """Map xlsx Affinity cell to GD enum token (case-insensitive)."""
    if card_name and card_name in UNIT_AFFINITY_OVERRIDES:
        raw = UNIT_AFFINITY_OVERRIDES[card_name]
    s = str(raw or "").strip()
    if not s:
        return None
    if s in AFFINITY:
        return AFFINITY[s]
    low = s.lower()
    if low in AFFINITY_WORDS:
        return AFFINITY_WORDS[low]
    for label, token in AFFINITY.items():
        if label.lower() == low:
            return token
    return None

RARITY = {
    "Common": "COMMON",
    "Uncommon": "UNCOMMON",
    "Rare": "RARE",
    "Legendary": "LEGENDARY",
    "Exotic": "EXOTIC",
    "1": "COMMON",
    "2": "UNCOMMON",
    "3": "RARE",
    "4": "LEGENDARY",
    "5": "EXOTIC",
}


def is_demo_card(card: dict) -> bool:
    demo = card.get("Demo")
    return demo is not None and str(demo).strip().lower() == "yes"


def load_workbook():
    # data_only=False keeps COUNTIF header formulas readable; card rows store plain "Yes"/None.
    return openpyxl.load_workbook(XLSX, data_only=False)


def sheet_rows(wb, name: str) -> tuple[list[str], list[dict]]:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else "" for h in rows[1]]
    cards = []
    for row in rows[2:]:
        if not row or not row[0]:
            continue
        card = {
            headers[i]: row[i]
            for i in range(len(headers))
            if headers[i] and i < len(row)
        }
        card["_name"] = str(row[0]).strip()
        cards.append(card)
    return headers, cards


def gd_name(name: str) -> str:
    return CARD_RENAMES.get(name, NAME_ALIASES.get(name, name))


def normalize_card_reference(name: str) -> str:
    return gd_name(name.strip())


def normalize_formula_text(formula: str) -> str:
    text = str(formula or "")
    for old, new in sorted(CARD_RENAMES.items(), key=lambda kv: -len(kv[0])):
        text = text.replace(old, new)
    return text


def parse_zone(text) -> list[tuple[int, int]]:
    if not text:
        return []
    cells = []
    for r, line in enumerate(str(text).strip().splitlines()):
        row = re.sub(r"\s+", "", line.strip())
        for c, ch in enumerate(row):
            if ch == "1":
                cells.append((r, c))
    return cells


def parse_summon_cost(formula: str) -> int | None:
    if not formula:
        return None
    m = re.findall(r"(\d+)\s*(?:crystals?|cost)\s*$", str(formula).strip(), re.I)
    if m:
        return int(m[-1])
    m2 = re.search(r"(\d+)\s*(?:crystals?|cost)", str(formula), re.I)
    return int(m2.group(1)) if m2 else None


def esc(s: str) -> str:
    return s.replace("\\", "\\\\").replace('"', '\\"')


def zone_to_gd(cells: list[tuple[int, int]]) -> str:
    if not cells:
        return "_z([])"
    parts = ", ".join(f"[{r},{c}]" for r, c in sorted(cells))
    return f"_z([{parts}])"


def _affinity_word(word: str) -> str | None:
    return AFFINITY_WORDS.get(word.strip().lower())


def _normalize_formula(formula: str) -> str:
    f = str(formula or "").strip().replace("\n", " ")
    f = f.replace("\u2019", "'").replace("\u2018", "'")
    f = re.sub(r"\s+", " ", f)
    f = re.sub(r"\s*\+\s*\d+\s*(?:crystals?|cost)\s*$", "", f, flags=re.I)
    return f.strip()


def _count_name_or_affinity_material(count: int, label: str) -> list[dict]:
    label = label.strip()
    aff = _affinity_word(label)
    if aff:
        return [{"affinity": aff}] * count
    return [{"name_contains": label.lower()}] * count


def _parse_or_label(label: str) -> dict | None:
    """Parse 'Arcane or Anima' / 'Elven or Elf' style labels."""
    if " or " not in label.lower():
        return None
    tokens = [t.strip() for t in re.split(r"\s+or\s+", label, flags=re.I) if t.strip()]
    if len(tokens) < 2:
        return None
    affs: list[str] = []
    names: list[str] = []
    for tok in tokens:
        tok = re.sub(r"\s+cards?$", "", tok, flags=re.I)
        aff = _affinity_word(tok)
        if aff:
            affs.append(aff)
        else:
            names.append(tok.lower())
    if affs and not names:
        return {"affinities": affs}
    if names and not affs:
        return {"name_contains_any": names}
    return None


def _parse_material_part(part: str) -> list[dict]:
    part = part.strip().rstrip(".")
    if not part or re.fullmatch(r"\d+\s*(?:crystals?|cost)\.?", part, re.I):
        return []

    low = part.lower()

    if "mutagen flag" in low:
        cond: dict = {"has_flag": "mutagen"}
        if "bio" in low:
            cond["affinity"] = "BIO"
        return [cond]

    if "princess flag" in low or "'princess'" in low or "princess" in low:
        return [{"name_contains": "princess"}]

    patterns: list[tuple[str, callable]] = [
        (
            r"(\d+)\s+(\w+)\s+\(def\s*≥?\s*(\d+)\)",
            lambda m: (
                [{"affinity": _affinity_word(m.group(2)), "min_def": int(m.group(3))}]
                * int(m.group(1))
                if _affinity_word(m.group(2))
                else []
            ),
        ),
        (
            r"(\d+)\s+(\w+)\s+\(≥?\s*(\d+)\s+def\)",
            lambda m: (
                [{"affinity": _affinity_word(m.group(2)), "min_def": int(m.group(3))}]
                * int(m.group(1))
                if _affinity_word(m.group(2))
                else []
            ),
        ),
        (
            r"1\s+tank\s+unit\s*\(≥?\s*(\d+)\s+cost\)",
            lambda m: [{"name_contains": "tank", "min_cost": int(m.group(1))}],
        ),
        (
            r"(\d+)\s+any\s+units?\s*\(≥?\s*(\d+)\s+def\)",
            lambda m: [{"min_def": int(m.group(2))}] * int(m.group(1)),
        ),
        (
            r"(\d+)\s+(\w+)\s+cards?\s*\(≥?\s*(\d+)\s+cost\)",
            lambda m: (
                [{"affinity": _affinity_word(m.group(2)), "min_cost": int(m.group(3))}]
                * int(m.group(1))
                if _affinity_word(m.group(2))
                else []
            ),
        ),
        (
            r"(\d+)\s+(\w+)\s+\(≥?\s*(\d+)\s+cost\)",
            lambda m: (
                [{"affinity": _affinity_word(m.group(2)), "min_cost": int(m.group(3))}]
                * int(m.group(1))
                if _affinity_word(m.group(2))
                else []
            ),
        ),
        (
            r"(\d+)\s+(\w+)\s+cards?\s*\((\d+)\s+cost\)",
            lambda m: (
                [{"affinity": _affinity_word(m.group(2)), "min_cost": int(m.group(3))}]
                * int(m.group(1))
                if _affinity_word(m.group(2))
                else []
            ),
        ),
        (
            r"(\d+)\s+(\w+)\s+cards?",
            lambda m: (
                [{"affinity": _affinity_word(m.group(2))}] * int(m.group(1))
                if _affinity_word(m.group(2))
                else _count_name_or_affinity_material(int(m.group(1)), m.group(2))
            ),
        ),
        (
            r"1\s+(\w+)\s+card\s*\(≥?\s*(\d+)\s+cost\)",
            lambda m: (
                [{"affinity": _affinity_word(m.group(1)), "min_cost": int(m.group(2))}]
                if _affinity_word(m.group(1))
                else []
            ),
        ),
        (
            r"1\s+(\w+)\s+card",
            lambda m: (
                [{"affinity": _affinity_word(m.group(1))}]
                if _affinity_word(m.group(1))
                else []
            ),
        ),
        (
            r"(\d+)\s+['']([^'']+)['']\s+name",
            lambda m: [{"name_contains": m.group(2).strip().lower().rstrip("s")}]
            * int(m.group(1)),
        ),
        (
            r"(\d+)\s+['']([^'']+)['']",
            lambda m: [{"name_contains": m.group(2).strip().lower().rstrip("s")}]
            * int(m.group(1)),
        ),
    ]

    for pat, builder in patterns:
        m = re.match(pat, part, re.I)
        if m:
            result = builder(m)
            if result:
                return result

    m = re.match(r"(\d+)\s+(.+?)\s+(?:cards?|units?)\s*$", part, re.I)
    if m:
        label = m.group(2).strip()
        or_cond = _parse_or_label(label)
        if or_cond:
            return [or_cond] * int(m.group(1))
        return _count_name_or_affinity_material(int(m.group(1)), label)

    m = re.match(r"1\s+(.+?\s+or\s+.+)$", part, re.I)
    if m:
        or_cond = _parse_or_label(m.group(1).strip())
        if or_cond:
            return [or_cond]

    m = re.match(r"(\d+)\s+(\w+)\s*$", part, re.I)
    if m:
        aff = _affinity_word(m.group(2))
        if aff:
            return [{"affinity": aff}] * int(m.group(1))

    if "armored" in low and "nature" in low:
        return [{"name_contains": "armored", "affinity": "NATURE"}]

    m = re.match(r"1\s+(.+)", part, re.I)
    if m:
        body = m.group(1).strip()
        body = re.sub(r"\s+cards?$", "", body, flags=re.I)
        body = re.sub(r"\s+unit$", "", body, flags=re.I)
        quoted = re.match(r"['']([^'']+)['']", body)
        if quoted:
            return [{"name_contains": quoted.group(1).strip().lower()}]
        words = body.split()
        if words:
            aff = _affinity_word(words[0])
            if aff and len(words) == 1:
                return [{"affinity": aff}]
            if aff and len(words) == 2 and words[1].lower() in ("card", "cards", "unit"):
                return [{"affinity": aff}]
        if len(words) >= 2:
            return [{"card_name": normalize_card_reference(body)}]
        return [{"name_contains": body.lower()}]

    aff = _affinity_word(part)
    if aff:
        return [{"affinity": aff}]

    if " or " in low:
        if "princess" in low:
            return [{"name_contains": "princess"}]

    if len(part.split()) >= 2:
        return [{"card_name": normalize_card_reference(part)}]

    return [{"name_contains": part.lower()}]


def parse_formula_conditions(formula: str) -> list[dict]:
    f = _normalize_formula(formula)
    if not f:
        return []
    conditions: list[dict] = []
    for part in f.split("+"):
        conditions.extend(_parse_material_part(part.strip()))
    return conditions


def cond_to_gd(cond: dict) -> str:
    parts: list[str] = []
    for key in sorted(cond.keys()):
        val = cond[key]
        if key == "affinity":
            parts.append(f'"affinity": A.{val}')
        elif key == "affinities" and isinstance(val, list):
            aff_parts = ", ".join(f"A.{a}" for a in val)
            parts.append(f'"affinities": [{aff_parts}]')
        elif key == "name_contains_any" and isinstance(val, list):
            name_parts = ", ".join(f'"{esc(str(n))}"' for n in val)
            parts.append(f'"name_contains_any": [{name_parts}]')
        elif isinstance(val, str):
            parts.append(f'"{key}": "{esc(val)}"')
        else:
            parts.append(f'"{key}": {val}')
    return "{" + ", ".join(parts) + "}"


def conds_to_gd(conds: list[dict], zone_size: int) -> str:
    if not conds:
        return f"_conds([], {zone_size})"
    inner = ", ".join(cond_to_gd(c) for c in conds)
    return f"_conds([{inner}], {zone_size})"


def sync_characters(text: str, units: list[dict]) -> tuple[str, int]:
    updated = 0
    for card in units:
        name = gd_name(card["_name"])
        aff = AFFINITY.get(str(card.get("Affinity", "")).strip(), None)
        if not aff:
            continue
        atk = int(card.get("ATK") or 0)
        def_ = int(card.get("DEF") or 0)
        cost = int(card.get("Cost") or 0)
        ability = str(card.get("Ability") or "").strip().replace("\n", " ")
        pat = (
            rf'(\t\t\["{re.escape(name)}",\s*CharacterData\.Affinity\.)\w+(,\s*)\d+(,\s*)\d+(,\s*)\d+(,)'
        )
        m = re.search(pat, text)
        if not m:
            continue
        repl = f'{m.group(1)}{aff}{m.group(2)}{atk}{m.group(3)}{def_}{m.group(4)}{cost}{m.group(5)}'
        text, n = re.subn(pat, repl, text, count=1)
        if n:
            updated += 1
        desc_pat = (
            rf'(\["{re.escape(name)}",[\s\S]*?CharacterData\.AbilityType\.\w+,\s*\{{[\s\S]*?\}},\s*\n\t\t\t")'
            rf'([^"]*)(")'
        )
        if ability and ability.lower() not in ("none", ""):
            text, n2 = re.subn(
                desc_pat,
                lambda mo: mo.group(1) + esc(ability) + mo.group(3),
                text,
                count=1,
            )
            if n2:
                updated += 1
    return text, updated


def sync_traps(text: str, traps: list[dict]) -> tuple[str, int]:
    updated = 0
    for card in traps:
        name = gd_name(card["_name"])
        cost = int(card.get("Cost") or 0)
        pat = rf'(\["{re.escape(name)}",\s*)\d+(\,\s*TrapData)'
        if re.search(pat, text):
            text, n = re.subn(pat, rf"\g<1>{cost}\g<2>", text, count=1)
            updated += n
        ability = str(card.get("Ability") or "").strip().replace("\n", " ")
        if ability:
            desc_pat = (
                rf'(\["{re.escape(name)}",[\s\S]*?TrapData\.TrapEffectType\.\w+,\s*'
                rf'\{{[\s\S]*?\}},\s*\n\t\t\t")([^"]*)(")'
            )
            text, n2 = re.subn(
                desc_pat,
                lambda mo: mo.group(1) + esc(ability) + mo.group(3),
                text,
                count=1,
            )
            if n2:
                updated += 1
    return text, updated


def sync_tech(text: str, techs: list[dict]) -> tuple[str, int]:
    updated = 0
    for card in techs:
        name = gd_name(card["_name"])
        cost = int(card.get("Cost") or 0)
        pat = rf'(\["{re.escape(name)}",\s*)\d+(\,\s*TechCardData)'
        if re.search(pat, text):
            text, n = re.subn(pat, rf"\g<1>{cost}\g<2>", text, count=1)
            updated += n
        ability = str(card.get("Ability") or "").strip().replace("\n", " ")
        if ability:
            desc_pat = (
                rf'(\["{re.escape(name)}",[\s\S]*?TechCardData\.TechEffectType\.\w+,\s*'
                rf'\{{[\s\S]*?\}},\s*"[^"]*",\s*\n\t\t\t")([^"]*)(")'
            )
            text, n2 = re.subn(
                desc_pat,
                lambda mo: mo.group(1) + esc(ability) + mo.group(3),
                text,
                count=1,
            )
            if n2:
                updated += 1
    return text, updated


def sync_unions(text: str, unions: list[dict]) -> tuple[str, int]:
    updated = 0
    for card in unions:
        name = gd_name(card["_name"])
        aff = AFFINITY.get(str(card.get("Affinity", "")).strip())
        if not aff:
            continue
        atk = int(card.get("ATK") or 0)
        def_ = int(card.get("DEF") or 0)
        full_ab = str(card.get("Full Ability") or "").strip().replace("\n", " ")
        part_ab = str(card.get("Partial Ability") or full_ab).strip().replace("\n", " ")
        full_f = normalize_formula_text(str(card.get("Full Formula") or "").strip().replace("\n", " "))
        part_f = normalize_formula_text(str(card.get("Partial Formula") or full_f).strip().replace("\n", " "))
        summon = parse_summon_cost(full_f)
        zone_cells = parse_zone(card.get("Union Zone"))

        header_pat = (
            rf'(_add\("{re.escape(name)}",\s*A\.)\w+,\s*(\d+),\s*(\d+),\s*(\d+),'
        )
        m = re.search(header_pat, text)
        if not m:
            continue
        summon_val = summon if summon is not None else int(m.group(4))
        repl_header = f'{m.group(1)}{aff}, {atk}, {def_}, {summon_val},'
        text, n = re.subn(header_pat, repl_header, text, count=1)
        if n:
            updated += 1

        if full_ab:
            ab_pat = (
                rf'(_add\("{re.escape(name)}"[\s\S]*?AB\.\w+,\s*\{{[\s\S]*?\}},\s*")'
                rf'([^"]*)("\s*,\s*")'
                rf'([^"]*)("\s*,\s*")'
                rf'([^"]*)("\s*,\s*")'
                rf'([^"]*)("\s*,\s*)'
            )
            text, n2 = re.subn(
                ab_pat,
                lambda mo: (
                    f'{mo.group(1)}{esc(full_ab)}{mo.group(3)}{esc(part_ab)}{mo.group(5)}'
                    f'{esc(full_f)}{mo.group(7)}{esc(part_f)}{mo.group(9)}'
                ),
                text,
                count=1,
            )
            if n2:
                updated += 1

        if zone_cells:
            zone_pat = rf'(_add\("{re.escape(name)}"[\s\S]*?)(\t\t_z\(\[[\s\S]*?\]\))'
            text, n3 = re.subn(
                zone_pat,
                lambda mo: mo.group(1) + "\t\t" + zone_to_gd(zone_cells),
                text,
                count=1,
            )
            if n3:
                updated += 1

        parsed_conds = parse_formula_conditions(full_f)
        if parsed_conds:
            zone_size = len(zone_cells) if zone_cells else max(len(parsed_conds), 5)
            gd_conds = conds_to_gd(parsed_conds, zone_size)
            cond_pat = (
                rf'(_add\("{re.escape(name)}"[\s\S]*?)(_conds\(\[[\s\S]*?\],\s*\d+\))'
            )
            text, n4 = re.subn(cond_pat, rf"\g<1>{gd_conds}", text, count=1)
            if n4:
                updated += 1

    return text, updated


def write_demo_flags(wb) -> int:
    flags = {}
    for sheet in ("Unit", "Trap", "Tech", "Union"):
        _, cards = sheet_rows(wb, sheet)
        for card in cards:
            demo = card.get("Demo")
            flags[gd_name(card["_name"])] = (
                demo is not None and str(demo).strip().lower() == "yes"
            )
    DEMO_FLAGS.parent.mkdir(parents=True, exist_ok=True)
    DEMO_FLAGS.write_text(json.dumps(flags, indent="\t") + "\n", encoding="utf-8")
    return sum(1 for v in flags.values() if v)


def _parse_booster_columns(headers: list[str]) -> dict[str, str]:
    """Map booster header string -> short pack name (e.g. 'Blightsilver')."""
    result: dict[str, str] = {}
    for h in headers:
        if not h:
            continue
        m = re.match(r"Booster Pack - (.+?) / \d+$", h)
        if m:
            result[h] = m.group(1)
            continue
        # Live COUNTIF header formulas, e.g. ="Booster Pack - Blightsilver / "&COUNTIF(...)
        for pack_name in ("Blightsilver", "Unseen Presence", "Wind of Dominance"):
            if pack_name in h:
                result[h] = pack_name
                break
    return result


def _default_weight(card: dict) -> int:
    rarity = card.get("Rarity")
    if rarity is not None:
        try:
            return RARITY_DEFAULT_WEIGHT.get(int(rarity), 100)
        except (TypeError, ValueError):
            pass
    cost = card.get("Cost")
    if cost is not None:
        try:
            if int(cost) >= 1000:
                return 5
        except (TypeError, ValueError):
            pass
    return 100


def _collect_booster_pools(wb) -> dict[str, dict[str, dict]]:
    """Return {pack_short_name: {card_name: {type, card_row}}}."""
    pools: dict[str, dict[str, dict]] = {}
    for sheet, default_type in SHEET_CARD_TYPE.items():
        headers, cards = sheet_rows(wb, sheet)
        booster_cols = _parse_booster_columns(headers)
        for header, short_name in booster_cols.items():
            pack_cards = pools.setdefault(short_name, {})
            for card in cards:
                val = card.get(header)
                if val is None or str(val).strip().lower() != "yes":
                    continue
                cname = gd_name(card["_name"])
                pack_cards[cname] = {
                    "type": default_type,
                    "row": card,
                }
    return pools


def _pack_unit_type_id(pack: dict) -> str:
    for entry in pack.get("card_pool", []):
        t = str(entry.get("card_type", "")).strip().lower()
        if t in ("unit", "character"):
            return t
    return "character"


def sync_custom_packs(wb) -> tuple[int, int, int]:
    if not CUSTOM_PACKS.exists():
        print("custom_packs.json not found — skipped booster sync")
        return 0, 0, 0

    pools = _collect_booster_pools(wb)
    packs: list = json.loads(CUSTOM_PACKS.read_text(encoding="utf-8"))
    updated_packs = 0
    added_total = 0
    removed_total = 0

    for pack in packs:
        short_name = str(pack.get("name", "")).replace("Booster Pack - ", "").strip()
        if short_name not in pools:
            continue

        desired = pools[short_name]
        old_pool: list = pack.get("card_pool", [])
        old_by_name = {str(e.get("card_name", "")): e for e in old_pool}
        unit_type = _pack_unit_type_id(pack)

        new_pool: list[dict] = []
        for cname in sorted(desired.keys()):
            meta = desired[cname]
            ctype = meta["type"]
            if ctype == "character":
                ctype = unit_type

            old_entry = old_by_name.get(cname)
            if old_entry is None:
                alias = WEIGHT_ALIASES.get(cname)
                if alias:
                    old_entry = old_by_name.get(alias)

            if old_entry is not None:
                weight = old_entry.get("weight", 100)
                ctype = str(old_entry.get("card_type", ctype))
            else:
                weight = _default_weight(meta["row"])

            new_pool.append(
                {
                    "card_name": cname,
                    "card_type": ctype,
                    "weight": weight,
                }
            )

        old_names = set(old_by_name.keys())
        new_names = set(desired.keys())
        added = new_names - old_names
        removed = old_names - new_names
        # Treat alias renames as updates, not add/remove churn
        for new_name, old_name in WEIGHT_ALIASES.items():
            if new_name in added and old_name in removed:
                added.discard(new_name)
                removed.discard(old_name)

        if new_pool != old_pool:
            pack["card_pool"] = new_pool
            updated_packs += 1
            added_total += len(added)
            removed_total += len(removed)

    CUSTOM_PACKS.write_text(json.dumps(packs, indent="\t") + "\n", encoding="utf-8")
    return updated_packs, added_total, removed_total


def main():
    wb = load_workbook()
    _, units = sheet_rows(wb, "Unit")
    _, traps = sheet_rows(wb, "Trap")
    _, techs = sheet_rows(wb, "Tech")
    _, unions = sheet_rows(wb, "Union")

    card_text = CARD_DB.read_text(encoding="utf-8")
    card_text, u1 = sync_characters(card_text, units)
    card_text, u2 = sync_traps(card_text, traps)
    card_text, u3 = sync_tech(card_text, techs)
    CARD_DB.write_text(card_text, encoding="utf-8")

    union_text = UNION_DB.read_text(encoding="utf-8")
    union_text, u4 = sync_unions(union_text, unions)
    UNION_DB.write_text(union_text, encoding="utf-8")

    demo_count = write_demo_flags(wb)
    pack_updates, pack_added, pack_removed = sync_custom_packs(wb)
    print(f"CardDatabase: {u1} character stat/desc touches, {u2} traps, {u3} tech (all in-game cards)")
    print(f"UnionDatabase: {u4} union touches (all in-game cards)")
    print(f"demo_flags.json: {demo_count} demo=yes entries")
    print(
        f"custom_packs.json: {pack_updates} pack(s) updated "
        f"(+{pack_added} cards, -{pack_removed} cards)"
    )


if __name__ == "__main__":
    main()
