#!/usr/bin/env python3
"""Clone card_data.xlsx with refined ability text + Comment column. Does not touch game code."""
from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "context" / "card_data.xlsx"
TS = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
OUT = ROOT / "context" / f"card_data_refine_{TS}.xlsx"

# --- Manual final text (user-confirmed design) ---
MANUAL: dict[str, dict[str, str]] = {
    "Plant-29": {
        "Ability": "Start of owner's turn: Flip a coin. Head: put Venom Flag on 1 exposed ally or foe card. Tail: put Mutagen Flag on any of your unit.",
        "comment": "Synced to card_data.xlsx Ability column.",
    },
    "Death Cobra": {
        "Ability": "At end of turn: choose 1 foe card. Put Venom flag on it.",
        "comment": "Synced to implemented rule; any foe card (Exposed or face-down flips up).",
    },
    "Dwarven Bomber": {
        "Ability": (
            "After battle: Flip 3 coins. Reveal up to that many foe cells (Heads). "
            "All Tails: destroy this card."
        ),
        "comment": "Fixed typos; confirmed 3-coin intent.",
    },
    "Armored Wolf": {
        "Ability": "Once in Reckoning: owner may grant this card +50 ATK and -50 DEF permanently.",
        "comment": "User confirmed -50 ATK was typo for -50 DEF.",
    },
    "Parasite Queen": {
        "Ability": (
            "After this card defends successfully, the attacker is not destroyed "
            "but cannot attack ever again."
        ),
        "comment": "User confirmed defender→attacker typo.",
    },
    "Electrogazer": {
        "Ability": "Disable all 0-cost traps on both fields.",
        "comment": "Confirmed: disable (not destroy/negate); both fields.",
    },
    "Sweet Lure Pod": {
        "Ability": "This turn, foe can only attack this card (if it survives).",
        "comment": "User confirmed targeting restriction while unit survives.",
    },
    "Ectoplasm": {
        "Ability": (
            "After this card is destroyed in Reckoning, revive 1 owned card (ability becomes None). "
            "Repeat the same battle once."
        ),
        "comment": "User confirmed repeat same battle once; grammar.",
    },
    "Nanami the Long Neck": {
        "Ability": "Once Exposed: foe chooses 1 Tech and shows it to the owner.",
        "comment": "Grammar fix; final design.",
    },
    "Doctopus": {
        "Ability": (
            "Once: when an ally Nature card is destroyed, owner may set this card's ATK to 0 "
            "to revive it at end of turn. Usable face-down."
        ),
        "comment": "Grammar shorten; no You/Your; final design.",
    },
    "Black Hole": {
        "Ability": "Destroy all Cosmic cards on foe's field. Discard all this player's Tech cards.",
        "comment": "Cosmo→Cosmic typo; no You/Your; shorten.",
    },
    "Blood Ritual": {
        "Ability": (
            "Destroy 1 card on this player's field without paying cost. "
            "Choose 1 Exposed foe unit. Its ATK&DEF becomes 0 permanently."
        ),
        "comment": "Preserve unit; No paid cost→without paying cost.",
    },
    "Welcoming Door": {
        "Ability": "Reveal any own cell.",
        "comment": "Shortened; cell terminology.",
    },
    "Trap Hole": {
        "Ability": "Flip 3 coins. Attacker loses 20 Crystals per Head.",
        "comment": "Grammar: coins plural; per Head wording.",
    },
    "Acid Trap Hole": {
        "Ability": "Flip 2 coins. Attacker loses 50 Crystals per Head.",
        "comment": "Grammar: coins plural; per Head wording.",
    },
    "Lava Trap Hole": {
        "Ability": "Flip 3 coins. Attacker loses 100 Crystals per Head.",
        "comment": "Grammar: coins plural; per Head wording.",
    },
    "Pepper Spray": {
        "Ability": (
            "Flip 2 coins. Heads: attacker loses 5 ATK per Head until attacker's next turn ends."
        ),
        "comment": "Grammar; removed double negative 'lose -5'.",
    },
    "Red Card": {
        "Ability": (
            "Flip 2 coins. Both Heads: that card cannot attack until attacker's next turn ends."
        ),
        "comment": "Fixed incomplete sentence.",
    },
    "Ancient Lich": {"Ability": "Immune to Tech.", "comment": "Shortened immunity wording."},
    "Araya the Eerie Dancer": {"Ability": "Immune to Tech.", "comment": "Shortened immunity wording."},
    "Goat Head": {"Ability": "Immune to traps.", "comment": "Shortened immunity wording."},
    "Tomb Bandit": {"Ability": "Immune to traps.", "comment": "Shortened immunity wording."},
    "Ironclad Sentinel": {
        "Ability": "Immune to 0-cost traps. Immune to Tech.",
        "comment": "Shortened immunity wording.",
    },
    "Xenospawn": {"Ability": "Immune to traps.", "comment": "Shortened immunity wording."},
    "Red Devil": {"Ability": "Immune to traps.", "comment": "Shortened immunity wording."},
    "Berserk": {
        "Ability": (
            "Choose 1 Exposed ally. It gets 1 extra attack but is this player's only attacker this turn. "
            "Cannot use if this player already attacked."
        ),
        "comment": "Shortened; Exposed terminology; no You/Your.",
    },
    "Arcane Duplication": {
        "Ability": (
            "Choose 1 card. Create a token copy on an empty cell. "
            "Destroy the token at end of this player's next turn."
        ),
        "comment": "Removed dangling 'This player.'; no You/Your; shortened.",
    },
    "Katana Strike": {
        "Ability": (
            "Choose 1 Exposed foe card. Destroy other Exposed cards in its column. "
            "Foe pays no cost."
        ),
        "comment": "Shortened; Exposed/cell/cost.",
    },
    "Rift Strike": {
        "Ability": (
            "Choose 1 Exposed foe card. Destroy other Exposed cards in its row. "
            "Foe pays no cost."
        ),
        "comment": "Shortened; Exposed/cell/cost.",
    },
    "Corrupted Spy": {
        "Ability": (
            "Reveal 3 foe cells. Pay 700 cost per trap or card found."
        ),
        "comment": "Shortened; cell/cost.",
    },
    "Accident": {
        "Ability": (
            "Destroy 1 Exposed foe card. If none, foe chooses 1 to Expose. Foe pays no cost."
        ),
        "comment": "Shortened; Exposed terminology.",
    },
    "Capnomancer": {
        "Full Ability": "At turn start: owner may destroy this card to revive Pyromancer.",
        "comment": "Confirmed: revive only; no coin flip.",
    },
    "Seraph Lawkeeper": {
        "Ability": (
            "Once: if a Divine or Anima card would be destroyed, revive it at end of turn. "
            "Usable face-down."
        ),
        "comment": "Grammar fix.",
    },
    "Wending the Wise Princess": {
        "Ability": (
            "Once Exposed: reveal 1 foe cell or put 1 Princess flag on an ally card."
        ),
        "comment": "Princess flag typo; Exposed terminology.",
    },
    "Angel Mauler": {
        "Ability": "In Reckoning, this card is not destroyed by Chaos cards.",
        "comment": "Confirmed: Reckoning-only (not traps/Tech).",
    },
    "Cloud Beast": {
        "Ability": "In Reckoning, this card is not destroyed by Anima cards.",
        "comment": "Confirmed: Reckoning-only (not traps/Tech).",
    },
    "Burning Phoenix": {
        "Full Ability": (
            "In Reckoning, this card is not destroyed by non-Union cards. "
            "If targeted by Tech, destroy this card."
        ),
        "comment": "Confirmed: non-Union immunity is units only; Tech still destroys.",
    },
    "Gemina the Supreme Queen": {
        "Full Ability": (
            "Both players cannot use Tech while this card is Exposed. Immune to traps."
        ),
        "comment": "Confirmed scope.",
    },
    "Wyriah the Roc": {
        "Full Ability": "Immune to trap effects.",
        "comment": "Shortened immunity wording.",
    },
    "Trash Penalty": {
        "Ability": "Foe chooses: clear all bluffs or lose 500 Crystals.",
        "comment": "Confirmed: damage = Crystals.",
    },
    "Pit Lord": {
        "Ability": (
            "Destroy this card after Reckoning with Divine card. "
            "After this card attacks, halve ATK&DEF permanently."
        ),
        "comment": "card→unit; Divine Unit→unit; grammar.",
    },
    "Lab Crawler": {
        "Ability": "With Mutagen flag: this card can target 3 cards.",
        "comment": "card→unit; Mutagen flag case.",
    },
    "Archbishop": {
        "Ability": (
            "If this card would be destroyed, owner may destroy 1 other Divine card "
            "on its side instead, even if face-down."
        ),
        "comment": "Confirmed: sacrifice target can be face-down.",
    },
    "Mountain Sage": {
        "Ability": "Double effect of Tech applied to this card.",
        "comment": "Tech card→Tech; card→unit.",
    },
    "Hyperspeed Saucer": {
        "Ability": "Gain +10 ATK&DEF permanently at end of each owner's turn.",
        "comment": "Shortened; grammar.",
    },
    "Railgun Tank": {
        "Ability": "After a successful attack, this card cannot attack on the owner's next turn.",
        "comment": "Shortened; card→unit.",
    },
    "War Genie ": {
        "Ability": "-10 ATK permanently after it attacks.",
        "comment": "Grammar: attacked→attacks.",
    },
    "Void Stalker": {
        "Ability": "+20 ATK if it attacks an Exposed card.",
        "comment": "Grammar: attack→attacks; card→unit.",
    },
    "Scout Probe": {
        "Ability": "After attacking: reveal 1 adjacent cell.",
        "comment": "Shortened structure.",
    },
    "Aether Warden": {
        "Ability": "In Reckoning: if it defends, attacker loses 300 Crystals.",
        "comment": "Added colon; shortened.",
    },
    "Fierce Gladiator": {
        "Ability": "Gain 500 Crystals on a successful defend.",
        "comment": "Crystal→Crystals amount wording.",
    },
    "Silver Dragon": {
        "Ability": "Pay 1000 cost per attack to command this card.",
        "comment": "Confirmed: pay each attack.",
    },
    "Lab Bloater": {
        "Ability": (
            "With Mutagen flag: owner can destroy both cards in Reckoning. "
            "Both players pay no cost."
        ),
        "comment": "Confirmed: both players pay no cost.",
    },
    "Melissa the Healer": {
        "Ability": "If the owner loses 500 or more Crystals, the owner recovers 300 Crystals.",
        "comment": "Grammar; shorthand ≥.",
    },
    "Joan the Faithful Warrior": {
        "Ability": "If at least 1 Exposed Divine card is on the field, this card gains 35 DEF.",
        "comment": "Shortened; exposed→Exposed; grammar.",
    },
    "Sniping Fairy": {
        "Ability": "After Exposed: -20 ATK at that turn end.",
        "comment": "Shortened; becomes exposed→After Exposed.",
    },
    "Mafia Associates": {
        "Ability": "End of Exposed turn: DEF becomes 0.",
        "comment": "Shortened; becomes→Exposed turn.",
    },
    "Cursed Well": {
        "Ability": "End of Exposed turn: +15 ATK.",
        "comment": "Shortened.",
    },
    "Hands in the Attic": {
        "Ability": "+10 ATK until Reckoning ends.",
        "comment": "Minor shorten.",
    },
    "Mine Guard": {
        "Ability": (
            "Prevent Miner or Mining card from being destroyed; destroy this card instead. "
            "Usable face-down."
        ),
        "comment": "card→unit; shortened quotes.",
    },
    "Grave Worm": {
        "Ability": "Each time foe loses Crystals: foe loses 20 more Crystals.",
        "comment": "Grammar lose→loses.",
    },
    "Sinister Cultist": {
        "Ability": "Whenever foe loses Crystals, lose 100 more Crystals.",
        "comment": "Grammar; shortened.",
    },
    "Mad Doctor": {
        "Ability": "+50 Crystals whenever any player uses a Tech.",
        "comment": "use Tech card→uses Tech; Crystal→Crystals.",
    },
    "Ore Transporter": {
        "Ability": "Whenever owner uses Tech, gain 20 Crystals.",
        "comment": "tech→Tech; receive→gain.",
    },
    "Magical Butterfly": {
        "Ability": "Whenever foe activates Tech, +10 ATK&DEF permanently. Usable face-down.",
        "comment": "tech card→Tech; shortened.",
    },
    "Ethereal Enchanter": {
        "Ability": "Whenever owner uses Tech, +50 ATK&DEF until foe turn ends.",
        "comment": "tech→Tech; shortened.",
    },
    "Greedy Gremlin": {
        "Ability": (
            "Whenever owner gains Crystals from any source, +10 ATK&DEF permanently (once)."
        ),
        "comment": "Grammar; shortened.",
    },
    "Corrupted Gremlin": {
        "Ability": (
            "Whenever owner gains Crystals from any source, +10 ATK permanently (max +50)."
        ),
        "comment": "Grammar gain→gains.",
    },
    "The Ancient One": {
        "Ability": "Foe cannot use Tech while this card is Exposed.",
        "comment": "Shortened; card→unit.",
    },
    "Thunder Daddy": {
        "Ability": "After Exposed: owner can only command this card to attack.",
        "comment": "Shortened; card→unit.",
    },
    "Gluttonous Amoeba": {
        "Ability": "In Reckoning: foe's next attack must target a face-down card.",
        "comment": "battles→In Reckoning; card→unit.",
    },
    "Lessor Leech": {
        "Ability": "Attacker loses 400 Crystals. Owner gains the same amount.",
        "comment": "Capitalization; grammar.",
    },
    "Sticky Grappler": {
        "Ability": "If this card defends, flip coin. Heads: foe turn ends immediately.",
        "comment": "card→unit; Head→Heads; shortened.",
    },
    "Long Tongue": {
        "Ability": "After a successful attack, flip a coin. Heads: gain 20 Crystals.",
        "comment": "Head→Heads; Crystal→Crystals.",
    },
    "Slug-11": {
        "Ability": "At end of the owner's turn, foe loses 300 Crystals per Mutagen flag on the field.",
        "comment": "Shortened; grammar.",
    },
    "Epsilon The Withered": {
        "Ability": (
            "In Reckoning: +20 ATK&DEF this turn; all ally cards -20 ATK&DEF this turn."
        ),
        "comment": "Shortened dual clause.",
    },
    "Moon Tribe Demolisher": {
        "Ability": (
            "Once in Reckoning: destroy defender. If no other ally Moon card, destroy this card."
        ),
        "comment": "battle calculation→Reckoning; card→unit.",
    },
    "Space Junk Janitor": {
        "Ability": (
            "Before Reckoning: owner may destroy 1 trap on its side for +20 ATK&DEF."
        ),
        "comment": "trap card→trap; shortened.",
    },
    "Quezil the Space Scavenger": {
        "Ability": "Once Exposed: gain 50 Crystals per card in void.",
        "comment": "Shortened.",
    },
    "Goliath the Wonder Seeker": {
        "Ability": "At end of the owner's turn, if this card survived, reveal 1 random foe cell.",
        "comment": "Shortened.",
    },
    "Hairpin Assassin": {
        "Ability": "In Reckoning: owner may pay 100 Crystals for +10 ATK.",
        "comment": "Crystal→Crystals; colon.",
    },
    "Leopard Jailer": {
        "Ability": "If this card attacks a card, target cannot attack until its next turn ends.",
        "comment": "card→unit; shortened.",
    },
    "Zealot": {
        "Ability": "If used for Union summon, Union card gets +40 ATK&DEF.",
        "comment": "Shortened; card→unit.",
    },
    "Evolving Cell": {
        "Ability": "Cannot target cards with 0 DEF. Once: if defends, drain 35 ATK&DEF.",
        "comment": "Shortened; shorthand DEF becomes 0.",
    },
    "Charm Mistress": {
        "Full Ability": "At start of foe's turn, flip a coin. Heads: foe can only attack once.",
        "comment": "Head→Heads; grammar.",
    },
    "Rebel King": {
        "Full Ability": "End of foe turn: owner selects 1 Exposed foe card; swap its ATK&DEF.",
        "comment": "Shortened; select→selects fixed in refine.",
    },
    "Scarlet Shroom": {
        "Full Ability": "Summoned: put Venom flag on all Exposed foe cards.",
        "comment": "Once Union→Summoned; card→unit.",
    },
    "Diamond Unicorn": {
        "Full Ability": "Summoned: +15 ATK until turn end.",
        "comment": "Once Union→Summoned; shortened.",
    },
    "Moon Tribe Shaman": {
        "Full Ability": "Summoned: revive 1 Moon non-Union card. Double its cost.",
        "comment": "Upon union→Summoned; card→unit.",
    },
    "Mass Extinction Dragon": {
        "Full Ability": (
            "Summoned: destroy all other Exposed ally units without paying cost."
        ),
        "comment": "Preserve units; No paid cost→without paying cost.",
    },
    "Acid Elemental": {
        "Full Ability": (
            "Summoned: put Acid flag on all Exposed foe cards. "
            "Acid flag: -10 DEF at end of foe's turn."
        ),
        "comment": "Summoned: prefix; card→unit; acid→Acid flag.",
    },
    "Magnet Elemental": {
        "Full Ability": (
            "Foe non-Arcane cards cannot target Exposed cards except this card."
        ),
        "comment": "cards→units; shortened.",
    },
    "Dwarven Drill": {
        "Full Ability": "If this card did not attack this turn, reveal 2 foe cells.",
        "comment": "Grammar; card→unit/cell.",
    },
    "Ice Elemental": {
        "Full Ability": "Battling card cannot attack until its next turn ends.",
        "comment": "Shortened; card→unit.",
    },
    "Thunder Elemental": {
        "Full Ability": "After successful attack, reveal 1 foe cell.",
        "comment": "card→cell; shortened.",
    },
    "Wood Elemental": {
        "Full Ability": "End of turn: +10 DEF permanently.",
        "comment": "Shortened.",
    },
    "Nuclear Elemental": {
        "Full Ability": "End of turn: +10 ATK permanently.",
        "comment": "Shortened.",
    },
    "Oblivion Dragon": {
        "Full Ability": "When defends: halve ATK; add that amount to DEF permanently.",
        "comment": "Shortened grammar.",
    },
    "Wild Manticore": {
        "Full Ability": "Gain +20 ATK at end of the owner's turn. Resets after attack.",
        "comment": "Shortened.",
    },
    "Mad Eye the Collector": {
        "Full Ability": (
            "In Reckoning: owner may destroy 1 trap on its side for +60 ATK&DEF."
        ),
        "comment": "trap card→trap; shortened.",
    },
    "Dark Champion": {
        "Full Ability": "Cannot target cards with 0 DEF. Destroy this card in Reckoning vs 0 DEF card.",
        "comment": "Shortened; fixed partial ?? placeholders in full.",
    },
    "Ceasefire": {
        "Ability": "Both players skip 1 turn (tax still applies).",
        "comment": "Shortened parenthetical.",
    },
    "Make Friend": {
        "Ability": (
            "Both players select 1 unit on the field. That unit cannot attack until "
            "this player's next turn ends."
        ),
        "comment": "monster→unit; grammar.",
    },
    "Arcane Nova": {
        "Ability": (
            "Destroy up to 3 Exposed foe cards (no cost). "
            "Discard all this player's Tech."
        ),
        "comment": "Shortened; Exposed; no cost→no cost.",
    },
    "Forest Fire": {
        "Ability": "Destroy all Nature cards on foe field. Discard all this player's Tech.",
        "comment": "cards→units; shortened.",
    },
    "Holy Arrival": {
        "Ability": "Destroy all Chaos cards on foe field. Discard all this player's Tech.",
        "comment": "cards→units; shortened.",
    },
    "Corrupted Heaven": {
        "Ability": "Destroy all Divine cards on foe field. Discard all this player's Tech.",
        "comment": "cards→units; shortened.",
    },
    "Earthquake": {
        "Ability": "Destroy all Anima cards on foe field. Discard all this player's Tech.",
        "comment": "cards→units; shortened.",
    },
    "Lab Destruction": {
        "Ability": "Destroy all Bio cards on foe field. Discard all this player's Tech.",
        "comment": "cards→units; shortened.",
    },
    "Anti-magic Field": {
        "Ability": "Destroy all Arcane cards on foe field. Discard all this player's Tech.",
        "comment": "cards→units; shortened.",
    },
    "Double Spy": {
        "Ability": "Requires Spy in void. Reveal 2 foe cells.",
        "comment": "Shortened condition.",
    },
    "Invisible Spy": {
        "Ability": "Requires Double Spy in void. Reveal 2 foe cells.",
        "comment": "Shortened condition.",
    },
    "Great Diplomacy": {
        "Ability": "Reveal up to 5 own cards.",
        "comment": "Select→Reveal; shortened.",
    },
    "Harsh Training": {
        "Ability": "+10 ATK permanently to 1 own card.",
        "comment": "Shortened.",
    },
    "Illegal Steroid": {
        "Ability": "+30 ATK to 1 card until turn end.",
        "comment": "Shortened.",
    },
    "Secret Dagger": {
        "Ability": "Face-down card only: Expose it; +15 ATK until turn end.",
        "comment": "Shortened; card→unit.",
    },
    "Bulletproof Vest ": {
        "Ability": "+15 DEF permanently to 1 Exposed card. On face-down: Expose it.",
        "comment": "Shortened; flip it up→Expose.",
    },
    "Siege Cannon": {
        "Ability": "This turn: once, destroy foe defending card in Reckoning.",
        "comment": "Shortened.",
    },
    "Silence Agent": {
        "Ability": "Destroy 1 Exposed foe card.",
        "comment": "card→unit.",
    },
    "Radar": {
        "Ability": "Reveal 3 foe cells.",
        "comment": "of→dropped.",
    },
    "Palace Party": {
        "Ability": "Reveal 1 own face-down card. Foe reveals 1 face-down card (if any).",
        "comment": "must→drops; shortened.",
    },
    "Essence Transfer": {
        "Ability": "Move all ATK&DEF modifiers from 1 card to another ally card.",
        "comment": "bonuses/debuffs→modifiers; shortened.",
    },
    "Time Travel ": {
        "Ability": "Once: revive 1 unit. Double its cost.",
        "comment": "Shortened revive wording; monster N/A.",
    },
    "Nobleman of Luminaso": {
        "Ability": "If this card battles and survives, pay 500 cost.",
        "comment": "card→unit; Crystal cost→cost.",
    },
    "Ethereal Marquees": {
        "Full Ability": (
            "-80 DEF if defends vs Arcane card. If battles and survives, pay 500 cost."
        ),
        "comment": "Shortened; Crystal cost→cost.",
    },
    "Armored Dino": {
        "Partial Ability": "In Reckoning, pay ??? cost to +??DEF",
        "Full Ability": "In Reckoning, pay 1000 cost to +60 DEF",
        "comment": "Crystal cost→cost; partial ??? kept.",
    },
    "Resurrection": {
        "Ability": "Once: revive 1 unit. ATK&DEF becomes 0; ability becomes None.",
        "comment": "Preserve unit; shortened shorthand.",
    },
    "Tech Copy": {
        "Ability": "Foe shows 1 Tech in hand. Copy it to this player's Tech Stack.",
        "comment": "Shortened.",
    },
    "Force Shield": {
        "Ability": "Select 1 card. It is not destroyed until foe turn ends.",
        "comment": "1card→1 card; shortened.",
    },
    "Wisp Light": {
        "Ability": "Destroy ally wisp cards; reveal that many foe cells.",
        "comment": "Shortened.",
    },
    "Lucky Day": {
        "Ability": "Flip coin. Heads: gain 600 Crystals. Always Heads if owner has ≤1000 Crystals.",
        "comment": "Head→Heads; shortened Always head clause.",
    },
    "Immortal Blood": {
        "Ability": "Revive 1 Immortal Vampire or Vampire Servant.",
        "comment": "Capitalization.",
    },
    "Yin Yang Swap ": {
        "Ability": "Swap ATK&DEF on 1 Exposed card until turn end.",
        "comment": "Shortened.",
    },
    "Casual Gambling": {
        "Ability": "Flip coin. Heads: gain 100 Crystals.",
        "comment": "Head→Heads; receive→gain.",
    },
    "Degen Gambling": {
        "Ability": "Flip coin. Heads: gain 200 Crystals. Tails: lose 400 Crystals.",
        "comment": "Head/Tail→Heads/Tails.",
    },
    "Hired Spanker": {
        "Ability": "Until foe turn ends: if foe attacks bluffed cell, gain 1000 Crystals.",
        "comment": "receive→gain; shortened.",
    },
    "Tax Avoidance": {
        "Ability": "End turn without paying tax (penalty still doubles).",
        "comment": "Shortened.",
    },
    "Compensation": {
        "Ability": "This turn: gain 20 Crystals per Dead End attacked.",
        "comment": "Shortened.",
    },
    "Phony Fight": {
        "Ability": "-10 DEF to 1 Exposed card until turn end. Foe gains 500 Crystals.",
        "comment": "receive→gains; shortened.",
    },
    "Borrow": {
        "Ability": "Gain 200 Crystals. End of turn: lose 300 Crystals.",
        "comment": "Shortened.",
    },
    "Loan": {
        "Ability": "Gain 600 Crystals. Next owner turn end: lose 800 Crystals.",
        "comment": "Shortened.",
    },
    "Mortgage": {
        "Ability": "Gain 1200 Crystals. End of turn: lose 1500 Crystals.",
        "comment": "Shortened.",
    },
    "Flimsy Axe": {
        "Ability": "Flip coin. Heads: +5 ATK to 1 ally card until turn end.",
        "comment": "Head→Heads.",
    },
    "Flimsy Shield": {
        "Ability": "Flip coin. Heads: +5 DEF to 1 ally card until turn end.",
        "comment": "Head→Heads.",
    },
    "All-out Tactics": {
        "Ability": "+10 ATK permanently to 1 ally card. DEF becomes 0 permanently.",
        "comment": "Shortened ATK&DEF becomes 0 style.",
    },
    "Fair Fight": {
        "Ability": "Both players select 1 card. +10 ATK&DEF permanently.",
        "comment": "Shortened.",
    },
    "Old Spear": {
        "Ability": "+5 ATK to 1 ally card until turn end.",
        "comment": "Shortened.",
    },
    "Old Shield": {
        "Ability": "+5 DEF to 1 ally card until turn end.",
        "comment": "Shortened.",
    },
    "Provoke": {
        "Ability": "End of foe turn: foe must attack selected cell first.",
        "comment": "Shortened.",
    },
    "Pretify": {
        "Ability": "+20 DEF permanently to ally or Exposed foe card. ATK=0 permanently.",
        "comment": "Shortened.",
    },
    "Stone Skin": {
        "Ability": "+20 DEF permanently to ally card. ATK=0 permanently.",
        "comment": "Shortened.",
    },
    "Battle Royale": {
        "Ability": "Both players select 1 card. +30 ATK&DEF.",
        "comment": "Shortened.",
    },
    "Guerrilla Tactics": {
        "Ability": (
            "Until foe turn ends: if foe attacks non-unit cell, flip coin. "
            "Heads: destroy attacking card."
        ),
        "comment": "Head→Heads; card→unit/cell.",
    },
    "Toll": {
        "Ability": "Until foe turn ends: foe pays 500 cost per attack.",
        "comment": "tax→cost; shortened.",
    },
    "Service Charge": {
        "Ability": "Until foe turn ends: foe pays 50 cost per attack.",
        "comment": "tax→cost; shortened.",
    },
    "Skull Amulet": {
        "Ability": "+5 ATK permanently to 1 Chaos card.",
        "comment": "card→unit.",
    },
    "Might": {
        "Ability": "+10 ATK to 1 ally Anima card until turn end.",
        "comment": "card→unit; shortened.",
    },
    "Arcane Knowledge": {
        "Ability": "+5 ATK&DEF to 1 ally Arcane card until foe turn ends.",
        "comment": "Shortened.",
    },
    "Air Shower": {
        "Ability": "Clear all flags on field.",
        "comment": "Shortened.",
    },
    "Mining Tax": {
        "Ability": "Until foe turn ends: attacking Dead End costs attacker 200 Crystals.",
        "comment": "Shortened conditional.",
    },
    "Ouija": {
        "Ability": "Reveal foe cells for each card in this player's void.",
        "comment": "Shortened.",
    },
    "Deep Mine": {
        "Ability": "Flip coin. Heads: reveal 2 foe cells.",
        "comment": "Head→Heads; square→cell.",
    },
    "Gold Panning": {
        "Ability": "Flip coin. Heads: reveal 1 foe cell.",
        "comment": "Head→Heads.",
    },
    "Flashlight": {
        "Ability": (
            "Reveal 1 foe column. Cannot attack this turn. "
            "Cannot use if already attacked."
        ),
        "comment": "Shortened.",
    },
    "Lantern": {
        "Ability": "Reveal 2 adjacent foe cells.",
        "comment": "Shortened.",
    },
    "GPS Tracker": {
        "Ability": "Reveal 2 adjacent cells next to Exposed foe card.",
        "comment": "Shortened.",
    },
    "Blessing": {
        "Ability": "Clear all flags from 1 ally card.",
        "comment": "Shortened.",
    },
    "Bright Sun": {
        "Ability": "Clear all flags from all cards on the field.",
        "comment": "Shortened.",
    },
    "Rugpull": {
        "Ability": "Both players lose 2000 Crystals. Discard all Tech in hand.",
        "comment": "Shortened.",
    },
    "Crack the Whip": {
        "Ability": (
            "Once: attacking card may attack twice. Cannot attack again until next owner turn end."
        ),
        "comment": "Shortened.",
    },
    "Full open fire": {
        "Ability": "Both players gain 2 extra attacks next turn.",
        "comment": "Shortened.",
    },
    "Lazy Scout": {
        "Ability": "Reveal 1 foe cell. Cannot attack. Cannot use if already attacked.",
        "comment": "Shortened.",
    },
    "Last Gambit": {
        "Ability": "If at most 5 cards are on the field, destroy 1 foe card.",
        "comment": "Grammar; card count.",
    },
    "Silence Zone": {
        "Ability": "Clear all bluffs. No bluffs for rest of duel.",
        "comment": "Shortened.",
    },
    "Big Mouth": {
        "Ability": "Select 1 cell. Bluff there becomes large permanently.",
        "comment": "Shortened.",
    },
    "Bare Fist Fight": {
        "Ability": "Both players discard 1 Tech from hand.",
        "comment": "Shortened.",
    },
    "Mine Detector": {
        "Ability": "Reveal 1 foe cell. If trap: destroy it; gain 200 Crystals.",
        "comment": "Shortened.",
    },
    "Natural Healing": {
        "Ability": "If all Exposed ally cards are Nature: gain 40 Crystals each.",
        "comment": "Shortened.",
    },
    "Third Eye": {
        "Ability": "If last Tech in hand: reveal 3 foe cells.",
        "comment": "Shortened.",
    },
    "Punching Bag": {
        "Ability": "1 ally card: ATK&DEF becomes 0 permanently. Cannot be destroyed until foe turn ends.",
        "comment": "Shortened.",
    },
    "Scattered Shot": {
        "Ability": "Flip 3 coins. Destroy Exposed foe cards for each Head. Foe pays no cost.",
        "comment": "Head→Heads; shortened.",
    },
    "Spotlight": {
        "Ability": "Flip 5 coins. Reveal foe cells for each Head.",
        "comment": "Shortened.",
    },
    "Tight Door": {
        "Ability": "Next turn: foe can attack only once.",
        "comment": "Shortened.",
    },
    "Mining Season": {
        "Ability": "Until foe turn ends: both players can only target face-down cards.",
        "comment": "both player→both players; card→unit.",
    },
    "Throne Duel": {
        "Ability": (
            "Until foe turn ends: both players can only attack cards with cost ≥800. "
            "Cannot use if already attacked."
        ),
        "comment": "Shorthand cost; grammar.",
    },
    "Joint Maneuver": {
        "Ability": (
            "Until foe turn ends: both players can only attack cards with cost ≤600. "
            "Cannot use if already attacked."
        ),
        "comment": "Shorthand cost.",
    },
    "Pillow Fight": {
        "Ability": (
            "Until foe turn ends: both players can only attack cards with cost ≤300. "
            "Cannot use if already attacked."
        ),
        "comment": "Shorthand cost.",
    },
    "Army Reformation": {
        "Ability": "Swap or relocate 1 card to any cell.",
        "comment": "Minor trim.",
    },
    "Institutionalization": {
        "Ability": "Target 1 foe card. Ability=None until foe next turn ends.",
        "comment": "Shortened.",
    },
    "Mana Drain": {
        "Ability": "Attacker loses 300 Crystals. Trapper gains that amount.",
        "comment": "Increase→gains; grammar.",
    },
    "Street Joke": {
        "Ability": "Reveal 1 trapper cell. Trapper gains 100 Crystals.",
        "comment": "Crystal→Crystals; receive→gains.",
    },
    "Anti-virus": {
        "Ability": (
            "If attacker is Bio: flip coin. Heads: destroy it. "
            "Always Heads if trapper has ≤1000 Crystals."
        ),
        "comment": "Head→Heads; shortened Always head.",
    },
    "Witch Hunt": {
        "Ability": (
            "If attacker is Arcane: flip coin. Heads: destroy it. "
            "Always Heads if trapper has ≤1000 Crystals."
        ),
        "comment": "Head→Heads; shortened.",
    },
    "Standoff": {
        "Ability": (
            "If attacker is Anima: flip coin. Heads: destroy it. "
            "Always Heads if trapper has ≤1000 Crystals."
        ),
        "comment": "Head→Heads; shortened.",
    },
    "Apple of Adam": {
        "Ability": (
            "If attacker is Divine: flip coin. Heads: destroy it. "
            "Always Heads if trapper has ≤1000 Crystals."
        ),
        "comment": "Head→Heads; shortened.",
    },
    "Purify": {
        "Ability": (
            "If attacker is Chaos: flip coin. Heads: destroy it. "
            "Always Heads if trapper has ≤1000 Crystals."
        ),
        "comment": "Head→Heads; shortened.",
    },
    "Solar Flare": {
        "Ability": (
            "If attacker is Cosmic: flip coin. Heads: destroy it. "
            "Always Heads if trapper has ≤1000 Crystals."
        ),
        "comment": "Head→Heads; shortened.",
    },
    "Hunting Season": {
        "Ability": (
            "If attacker is Nature: flip coin. Heads: destroy it. "
            "Always Heads if trapper has ≤1000 Crystals."
        ),
        "comment": "Head→Heads; shortened.",
    },
    "Dart Trap": {
        "Ability": "If attacker is first attack this turn, destroy it.",
        "comment": "monster→attacker; shortened.",
    },
    "Science Cage": {
        "Ability": "If attacker is Bio: end turn immediately. Clear all flags on it.",
        "comment": "Is→is; shortened.",
    },
    "Loop Hole": {
        "Ability": "If attacker is Arcane: end turn immediately.",
        "comment": "Shortened.",
    },
    "Talisman of Light": {
        "Ability": "If attacker is Chaos: -50 ATK permanently.",
        "comment": "monster→attacker; shortened.",
    },
    "Forbidden Grail": {
        "Ability": "If attacker is Divine: -50 DEF permanently.",
        "comment": "monster→attacker; shortened.",
    },
    "Electric Fence": {
        "Ability": "If attacker is Nature: cannot attack until next turn ends.",
        "comment": "Shortened.",
    },
    "Discourage": {
        "Ability": (
            "If attacker is Anima: -15 ATK to all Exposed attacker cards until next turn ends."
        ),
        "comment": "cards→units; shortened.",
    },
    "Galaxy Toll": {
        "Ability": "If attacker is Cosmic: foe loses 1000 Crystals.",
        "comment": "lose→loses.",
    },
    "Union Cage": {
        "Ability": "If attacker is Union: cannot attack until next turn ends.",
        "comment": "card→Union; shortened.",
    },
    "Plunder": {
        "Ability": "Attacker chooses: trapper gains 2000 Crystals or destroy attacking unit.",
        "comment": "monster→unit; shortened.",
    },
    "Decoy Puppet": {
        "Ability": "This turn: attacker cannot attack with cards with cost ≤400.",
        "comment": "Shorthand cost.",
    },
    "Hard Scale": {
        "Ability": "Trapper units +5 DEF in Reckoning until attacker next turn ends.",
        "comment": "Shortened.",
    },
    "Steel Scale": {
        "Ability": "Trapper units +15 DEF in Reckoning until attacker next turn ends.",
        "comment": "Shortened.",
    },
    "Trick Door": {
        "Ability": "Trapper selects 1 card; swap/relocate to any cell. Repeat Reckoning if this cell.",
        "comment": "Shortened.",
    },
    "Boiling Oil": {
        "Ability": "All attacker cards -10 ATK in Reckoning until turn end.",
        "comment": "Shortened.",
    },
    "Lightning Rod": {
        "Ability": "If foe targets this trap with Tech: foe loses 1000 Crystals.",
        "comment": "card→trap; tech→Tech.",
    },
    "Dreamcatcher": {
        "Ability": "When foe uses Tech: trigger immediately; cancel that Tech.",
        "comment": "Confirmed: triggers when foe uses any Tech.",
    },
    "Stumble": {
        "Ability": "Attacker ends turn. Remaining attacks carry to next turn.",
        "comment": "Shortened.",
    },
    "Ruckus": {
        "Ability": "Until trapper turn ends: Exposed Bio units gain +5 ATK.",
        "comment": "monster→unit; shortened.",
    },
    "Counter Punch": {
        "Ability": "Attacker -5 DEF until trapper turn ends.",
        "comment": "get→drops; shortened.",
    },
    "Counter Tackle": {
        "Ability": "Attacker -15 DEF until trapper turn ends.",
        "comment": "Shortened.",
    },
    "Protective Instinct": {
        "Ability": "Trapper selects 1 Nature card. Until trapper turn ends: +10 ATK.",
        "comment": "card→unit; shortened.",
    },
    "Self-destruct": {
        "Ability": "Trapper selects 1 own card. +10 ATK until trapper turn ends; destroy it. No cost.",
        "comment": "pay no cost→No cost.",
    },
    "Bounty": {
        "Ability": (
            "Until trapper turn ends: if trapper destroys that card, "
            "gain Crystals equal to half of card cost."
        ),
        "comment": "Shortened.",
    },
    "Share the Pain": {
        "Ability": (
            "If trapper has ≤1000 Crystals: drain attacker Crystals until both players equal."
        ),
        "comment": "Shortened.",
    },
    "Tripwire": {
        "Ability": "Destroy attacker with ATK ≤20. Attacker pays no cost.",
        "comment": "with 20 or less→ATK ≤20.",
    },
    "Potent Poison": {
        "Ability": (
            "Attacker permanently -5 ATK&DEF. -5 more per trap in trapper void."
        ),
        "comment": "Grammar; shortened.",
    },
    "Soul Blast": {
        "Ability": "Attacker loses 150 Crystals per card in trapper void.",
        "comment": "lose→loses; cards→units.",
    },
    "Grudge": {
        "Ability": "Attacker permanently -10 ATK per card in trapper void.",
        "comment": "monster→attacker; shortened.",
    },
    "Flame Trap": {
        "Ability": "Attacker permanently -10 ATK.",
        "comment": "get→drops; shortened.",
    },
    "Hostage": {
        "Ability": "Trapper reveals 1 own cell. This turn: attacker cannot target that cell.",
        "comment": "Grammar; shortened.",
    },
    "Bait": {
        "Ability": "Trapper chooses 1 other own cell and reveals it.",
        "comment": "choose→chooses; grammar.",
    },
    "Blackmail": {
        "Ability": "Attacker chooses: discard 1 Tech or end turn immediately.",
        "comment": "Shortened.",
    },
    "Cursed Reflection": {
        "Ability": "Swap attacker ATK&DEF until trapper turn ends.",
        "comment": "Shortened.",
    },
    "Explosive Barrels": {
        "Ability": "Destroy attacking card. Trapper pays same cost as attacker.",
        "comment": "also pay→pays; grammar.",
    },
    "Hypnosis ": {
        "Ability": "Attacking card cannot attack next turn.",
        "comment": "Shortened.",
    },
    "Echo Barrier": {
        "Ability": "This turn: attacker cannot attack again.",
        "comment": "Shortened.",
    },
    "Defensive Pheromone": {
        "Ability": "Trapper swaps 1 Armored Nature card with this cell. Repeat Reckoning.",
        "comment": "switch→swaps; quotes trimmed.",
    },
    "Snare Trap": {
        "Ability": "Attacker card ability becomes None until next turn ends.",
        "comment": "Shortened.",
    },
    "Brainwash": {
        "Ability": "Attacker chooses own ally as attack target.",
        "comment": "Shortened.",
    },
    "Bunker": {
        "Ability": "This turn: attacker cannot target surrounding cells.",
        "comment": "Shortened.",
    },
    "Foul Gas": {
        "Ability": "All attacker cards -5 ATK until turn end.",
        "comment": "Shortened.",
    },
    "Bingo the Chrono Rabbit": {
        "Ability": "Once: when this card defends, the attack does nothing.",
        "comment": "Grammar fix.",
    },
    "Venom Toad": {
        "Ability": (
            "After Reckoning: Put Venom flag on foe card. "
            "In Reckoning: destroy foe cards with Venom flag."
        ),
        "comment": "Confirmed: destroy-on-Venom applies every Reckoning.",
    },
    "Night Whisperer": {
        "Ability": "+30 ATK&DEF per Exposed card with wisp in name on its side.",
        "comment": "Confirmed: counts units whose name contains wisp.",
    },
}


def _reasons(old: str, new: str) -> list[str]:
    reasons: list[str] = []
    if old == new:
        return reasons
    if old.strip().lower() in ("none", "") and new.strip().lower() in ("none", ""):
        return reasons
    pairs = [
        (r"ATK/DEF", "ATK/DEF→ATK&DEF"),
        (r"ATK and DEF", "ATK and DEF→ATK&DEF"),
        (r"Affinity", "dropped redundant Affinity"),
        (r"face-up|face up", "face-up→Exposed"),
        (r"\bexposed\b", "exposed→Exposed"),
        (r"square", "square→cell"),
        (r"Flip \d coin[^s]", "Flip N coin→coins"),
        (r"Head:|If head|if head", "Head→Heads"),
        (r"Tail:|If tail|if tail", "Tail→Tails"),
        (r"both player", "both player→both players"),
        (r"foe lose|foe choose|attacker have|owner gain|owner use|owner have", "subject-verb grammar"),
        (r"unaffected by|not affected by|not destroyed by", "shortened immunity wording"),
        (r"Tech card|tech card", "Tech capitalization"),
        (r"\b(you|your)\b", "removed You/Your"),
        (r"Venom Flag|Mutagen Flag|princess flag|princes flag", "flag sentence case"),
        (r"0 cost", "0 cost→0-cost"),
        (r"\bmonster\b", "monster→unit"),
        (r"\bAt the end of this turn\b", "End of turn:"),
        (r"\bAt battle calculation\b", "In Reckoning:"),
        (r"\bOnce Union\b", "Once Union→Summoned:"),
    ]
    for pat, msg in pairs:
        if re.search(pat, old, re.I) and (not re.search(pat, new, re.I) or msg.startswith("dropped")):
            if msg not in reasons:
                reasons.append(msg)
    if len(new) < len(old) - 15 and "shortened" not in str(reasons):
        reasons.append("shortened")
    if not reasons:
        reasons.append("wording polish")
    return reasons


def apply_shorthand_symbols(t: str) -> str:
    """Prefer ATK&DEF and comparison symbols (≥ ≤ =) over spelled-out math."""
    t = re.sub(r"\bATK and DEF\b", "ATK&DEF", t, flags=re.I)
    t = re.sub(r"\bATK/DEF\b", "ATK&DEF", t, flags=re.I)
    # Cost / count comparisons — specific patterns before generic N or more/less
    t = re.sub(
        r"\bcan only attack with unit with (\d+) or more cost\b",
        r"can only attack units with cost ≥\1",
        t,
        flags=re.I,
    )
    t = re.sub(
        r"\bcan only attack with unit with (\d+) or less cost\b",
        r"can only attack units with cost ≤\1",
        t,
        flags=re.I,
    )
    t = re.sub(r"\bunits costing (\d+) or more cost\b", r"units with cost ≥\1", t, flags=re.I)
    t = re.sub(r"\bunits costing (\d+) or less cost\b", r"units with cost ≤\1", t, flags=re.I)
    t = re.sub(r"\bunits costing (\d+) or more\b", r"units with cost ≥\1", t, flags=re.I)
    t = re.sub(r"\bunits costing (\d+) or less\b", r"units with cost ≤\1", t, flags=re.I)
    t = re.sub(r"\bcards with cost (\d+) or more\b", r"cards with cost ≥\1", t, flags=re.I)
    t = re.sub(r"\bcards with cost (\d+) or less\b", r"cards with cost ≤\1", t, flags=re.I)
    t = re.sub(r"\b(\d+) or more Crystals\b", r"≥\1 Crystals", t, flags=re.I)
    t = re.sub(r"\b(\d+) or less Crystals\b", r"≤\1 Crystals", t, flags=re.I)
    t = re.sub(r"\b(\d+) or more Crystal\b", r"≥\1 Crystals", t, flags=re.I)
    t = re.sub(r"\b(\d+) or less Crystal\b", r"≤\1 Crystals", t, flags=re.I)
    t = re.sub(r"\b(\d+) or more cells\b", r"at least \1 cells", t, flags=re.I)
    t = re.sub(r"\b(\d+) or less cells\b", r"at most \1 cells", t, flags=re.I)
    t = re.sub(r"\b(\d+) or less\b", r"≤\1", t, flags=re.I)
    t = re.sub(r"\b(\d+) or more\b", r"≥\1", t, flags=re.I)
    t = re.sub(r"\bequal to number of heads?\b", "for each Head", t, flags=re.I)
    t = re.sub(r"\bOnce Union,\s*", "Summoned: ", t, flags=re.I)
    t = re.sub(r"\bUpon union,\s*", "Summoned: ", t, flags=re.I)
    t = re.sub(r"\bOn Union:\s*", "Summoned: ", t, flags=re.I)
    t = re.sub(r"\bCrystal cost\b", "cost", t, flags=re.I)
    return t


def retain_card_terminology(t: str) -> str:
    """Preserve card/trap; convert monster→unit; do not cross-rename card↔unit."""
    t = t.replace("non-unit", "\x00NU\x00")
    t = re.sub(r"\bmonsters\b", "units", t, flags=re.I)
    t = re.sub(r"\bmonster\b", "unit", t, flags=re.I)
    return t.replace("\x00NU\x00", "non-unit")


def finalize_ability_text(t: str) -> str:
    """Apply final style rules after all other transforms."""
    if not t or str(t).strip().lower() == "none":
        return t
    # :: → : and fix :,
    t = re.sub(r":{2,}", ":", t)
    t = re.sub(r":\s*,", ":", t)
    t = retain_card_terminology(t)
    t = apply_grammar_polish(t)
    # Expand banned shorthands
    t = re.sub(r"\s*=\s*#\s*Heads?\b", " for each Head", t, flags=re.I)
    t = re.sub(r"\s*=\s*#\s*(\w+)", r" for each \1", t, flags=re.I)
    t = re.sub(r"≥1\b", "at least 1", t)
    t = re.sub(r"≥(\d+) cells\b", r"at least \1 cells", t)
    t = re.sub(r"≤(\d+) cards\b", r"at most \1 cards", t)
    t = re.sub(r"\(face-down OK\)", "even if face-down", t, flags=re.I)
    t = re.sub(r"\bface-down OK\b", "even if face-down", t, flags=re.I)
    t = re.sub(r"= ½\b", "half of", t)
    t = re.sub(r"ATK&DEF=0", "ATK&DEF becomes 0", t, flags=re.I)
    t = re.sub(r"ATK=0", "ATK becomes 0", t, flags=re.I)
    t = re.sub(r"DEF=0", "DEF becomes 0", t, flags=re.I)
    t = re.sub(r"ability=None", "ability becomes None", t, flags=re.I)
    t = re.sub(r"\brevive ([^.]+?) from void\b", r"revive \1", t, flags=re.I)
    t = re.sub(r"\+ATK&DEF = ½\b", "+ATK&DEF equal to half of", t, flags=re.I)
    t = re.sub(r"\+= ½\b", "equal to half of", t, flags=re.I)
    t = re.sub(r"  +", " ", t).strip()
    if t:
        t = t[0].upper() + t[1:]
    return t


def apply_grammar_shorten(t: str) -> str:
    t = re.sub(r"\bit attack\b", "it attacks", t, flags=re.I)
    t = re.sub(r"\bit gain\b", "it gains", t, flags=re.I)
    t = re.sub(r"\bit lose\b", "it loses", t, flags=re.I)
    t = re.sub(r"\bit perform\b", "it performs", t, flags=re.I)
    t = re.sub(r"\bit get\b", "it gets", t, flags=re.I)
    t = re.sub(r"\bit have\b", "it has", t, flags=re.I)
    t = re.sub(r"\bowner lose\b", "owner loses", t, flags=re.I)
    t = re.sub(r"\bowner gain\b", "owner gains", t, flags=re.I)
    t = re.sub(r"\bthis card gain\b", "this card gains", t, flags=re.I)
    t = re.sub(r"\bwas commanded\b", "is commanded", t, flags=re.I)
    t = re.sub(r"\bAt the end of this turn\b", "At end of turn:", t, flags=re.I)
    t = re.sub(r"\bAt the end of foe's turn\b", "At end of foe's turn:", t, flags=re.I)
    t = re.sub(r"\bAt the start of each turn\b", "At each turn start:", t, flags=re.I)
    t = re.sub(r"\bStart of their turn\b", "At turn start:", t, flags=re.I)
    t = re.sub(r"\bOwner's turn starts\b", "At turn start:", t, flags=re.I)
    t = re.sub(r"\bAt battle calculation\b", "In Reckoning:", t, flags=re.I)
    t = re.sub(r"\bduring battle calculation\b", "in Reckoning", t, flags=re.I)
    t = re.sub(r"\bAfter successfully attack\b", "After a successful attack", t, flags=re.I)
    t = re.sub(r"\bAfter successful attack\b", "After a successful attack", t, flags=re.I)
    t = re.sub(r"\bAfter performed an attack\b", "After attacking", t, flags=re.I)
    t = re.sub(r"\bAfter this card battled\b", "After battling", t, flags=re.I)
    t = re.sub(r"\bAfter it is destroy\b", "After this card is destroyed", t, flags=re.I)
    t = re.sub(r"\bAfter destroyed\b", "After this card is destroyed", t, flags=re.I)
    t = re.sub(r"\bis destroy\b", "is destroyed", t, flags=re.I)
    t = re.sub(r"\bwill be destroy\b", "would be destroyed", t, flags=re.I)
    t = re.sub(r"\brevived that destroyed card\b", "revive that card", t, flags=re.I)
    t = re.sub(r"\bCannot be destroyed by non-union cards\b", "Immune to non-Union cards", t, flags=re.I)
    t = re.sub(r"\bCannot be destroyed by (\w+)\b", r"Immune to \1", t, flags=re.I)
    t = re.sub(r"\bunaffected by trap effect\b", "Immune to trap effects", t, flags=re.I)
    t = re.sub(r"\bDisable all 0-cost traps\b", "Disable all 0-cost traps", t, flags=re.I)
    t = re.sub(r"\bon both player's field\b", "on both fields", t, flags=re.I)
    t = re.sub(r"\bon both players's field\b", "on both fields", t, flags=re.I)
    t = re.sub(r"\bput 1 venom flag\b", "Put Venom flag", t, flags=re.I)
    t = re.sub(r"\badd venom flag to\b", "Put Venom flag on", t, flags=re.I)
    t = re.sub(r"\bPut Venom flag to\b", "Put Venom flag on", t, flags=re.I)
    t = re.sub(r"\bput venom flag to\b", "Put Venom flag on", t, flags=re.I)
    t = re.sub(r"\bvenom flag\b", "Venom flag", t, flags=re.I)
    t = re.sub(r"\bmutagen flag\b", "Mutagen flag", t, flags=re.I)
    t = re.sub(r"\bWithout Mutagen Flag\b", "Without Mutagen flag", t)
    t = re.sub(r"\bWith Mutagen Flag\b", "With Mutagen flag", t)
    t = re.sub(r"\breceive (\d+) Crystal\b", r"gain \1 Crystals", t, flags=re.I)
    t = re.sub(r"\bReceive (\d+) Crystal\b", r"Gain \1 Crystals", t)
    t = re.sub(r"\blose -(\d+)\b", r"-\1", t, flags=re.I)
    t = re.sub(r"\b  +", " ", t)
    return t.strip()


def apply_grammar_polish(t: str) -> str:
    """Short phrasing that stays grammatically correct."""
    reps: list[tuple[str, str]] = [
        (r"\bAfter a successful attack, cannot attack next owner turn\b",
         "After a successful attack, this card cannot attack on the owner's next turn"),
        (r"\bAfter a successful attack, cannot attack on owner's next turn\b",
         "After a successful attack, this card cannot attack on the owner's next turn"),
        (r"\bAfter attack:\b", "After attacking:"),
        (r"\bAfter attack,\b", "After attacking,"),
        (r"\bflip coin\b", "flip a coin"),
        (r"\bFlip coin\b", "Flip a coin"),
        (r"\bIf at least 1 Exposed Divine card on field\b",
         "If at least 1 Exposed Divine card is on the field"),
        (r"\bIf at least 1 Exposed (.+?) on field\b", r"If at least 1 Exposed \1 is on the field"),
        (r"\bOnce: if survived Reckoning\b", "Once: if this card survives Reckoning"),
        (r"\bIn Reckoning: not destroyed by non-Union cards\b",
         "In Reckoning, this card is not destroyed by non-Union cards"),
        (r"\bIn Reckoning: not destroyed by (\w+) cards\b",
         r"In Reckoning, this card is not destroyed by \1 cards"),
        (r"\bshow it to owner\b", "show it to the owner"),
        (r"\bdestroy foe card\b", "destroy 1 foe card"),
        (r"\bEnd of owner turn:", "At end of the owner's turn,"),
        (r"\bEnd of foe turn:", "At end of foe's turn,"),
        (r"\bEnd of turn:", "At end of turn,"),
        (r"\bEnd of Exposed turn:", "At end of Exposed turn,"),
        (r"\bat end of owner turn\b", "at end of the owner's turn"),
        (r"\bat end of foe turn\b", "at end of foe's turn"),
        (r"\beach owner turn\b", "each owner's turn"),
        (r"\bFoe turn start:\b", "At start of foe's turn,"),
        (r"\bOwner turn start:\b", "At start of owner's turn,"),
        (r"\bEach turn start:\b", "At each turn start:"),
        (r"\bTurn start:\b", "At turn start:"),
        (r"\bfoe turn ends immediately\b", "foe's turn ends immediately"),
        (r"\buntil foe turn ends\b", "until foe's turn ends"),
        (r"\bUntil foe turn ends\b", "Until foe's turn ends"),
        (r"\bfoe turn end\b", "foe's turn ends"),
        (r"\bowner turn end\b", "owner's turn ends"),
        (r"\bthe attacker attack is negated\b", "the attack does nothing"),
        (r"\battacker attack is negated\b", "the attack does nothing"),
        (r"\bHeads: negate attack\b", "Heads: the attack does nothing"),
        (r"\bnegate attack\b", "the attack does nothing"),
        (r"\bUsable face-down\b", "Usable while face-down"),
        (r"\bDrain (\d+)", r"drain \1"),
        (r"\+500 Crystals on successful defend\b", "Gain 500 Crystals on a successful defend"),
        (r"\bif a exposed\b", "if an Exposed"),
        (r"\bAttacker card ability\b", "The attacker's card ability"),
        (r"\bpay (\d+) Crystals for \+(\d+) ATK\b", r"pay \1 Crystals to gain +\2 ATK"),
        (r"\bIf this card attacks a card,\b", "If this card attacks, the target"),
        (r"\bOnce: after destroyed a card\b", "Once: after destroying a card,"),
        (r"\bafter attacked a\b", "after attacking a"),
        (r"\bAfter it performed attack\b", "After it attacks"),
        (r"\bPrevent Miner or Mining card\b", "Prevent a Miner or Mining card"),
        (r"\bReveal foe cells for each card in\b", "Reveal 1 foe cell for each card in"),
        (r"\bDestroy Exposed foe cards for each Head\b",
         "Destroy 1 Exposed foe card for each Head"),
        (r"\bWhenever foe loses Crystals, lose 100\b",
         "Whenever foe loses Crystals, foe loses 100"),
        (r"\bif survived,\b", "if this card survived,"),
        (r"\bBattling card cannot\b", "A battling card cannot"),
        (r"\bIn Reckoning if it defends\b", "In Reckoning, if it defends"),
        (r"\bon field\b", "on the field"),
        (r"\bper card in void\b", "per card in the void"),
        (r"\bAfter Exposed: -20 ATK at that turn end\b",
         "After becoming Exposed: -20 ATK at that turn's end"),
        (r"\bAfter Exposed:\b", "After becoming Exposed:"),
        (r"\bowner can only command this card to attack\b",
         "owner can only command this card to attack"),
        (r"\bIn Reckoning with non-Divine, Drain\b",
         "In Reckoning with non-Divine, drain"),
        (r"\bChoose 1 card, flip a coin\b", "Choose 1 card and flip a coin"),
        (r"\bAt turn start: Choose 1 card and flip a coin\b",
         "At turn start: choose 1 card and flip a coin"),
        (r"\bIf owner loses 500 or more Crystals, recover 300 Crystals\b",
         "If the owner loses 500 or more Crystals, the owner recovers 300 Crystals"),
        (r"\bOnce: when this card defends, the attack does nothing\b",
         "Once: when this card defends, the attack does nothing"),
        (r"\bOnce: when this card defends, the the attack\b",
         "Once: when this card defends, the attack"),
        (r"\bthe the attack\b", "the attack"),
        (r"\bIn Reckoning: owner can pay 2000 Crystals to destroy 1 foe card\b",
         "In Reckoning, the owner may pay 2000 Crystals to destroy 1 foe card"),
        (r"\bIn Reckoning: owner can pay 2000 Crystals to destroy foe card\b",
         "In Reckoning, the owner may pay 2000 Crystals to destroy 1 foe card"),
        (r"\bIn Reckoning: owner may pay 100 Crystals to gain \+10 ATK\b",
         "In Reckoning, the owner may pay 100 Crystals to gain +10 ATK"),
        (r"\bAt each turn start: put\b", "At each turn start, put"),
        (r"\bAt turn start: choose\b", "At turn start, choose"),
        (r"\bAt end of turn: choose\b", "At end of turn, choose"),
        (r"\bOnce: after destroying a card, this card can attack 1 more time\b",
         "Once: after destroying a card, this card can attack 1 more time"),
        (r"\bOnce: after destroying a card,\s*this card\b",
         "Once: after destroying a card, this card"),
        (r"\+10 ATK&DEF permanently at end of each owner turn\b",
         "Gain +10 ATK&DEF permanently at end of each owner's turn"),
        (r"or more of cells on its side is revealed\b",
         "or more cells on its side are revealed"),
        (r"if at least (\d+) cells on its side is revealed\b",
         r"if at least \1 cells on its side are revealed"),
        (r"if ≥(\d+) cells on its side is revealed\b",
         r"if at least \1 cells on its side are revealed"),
        (r"\bowner loses 300 Crystals per Mutagen flag on field\b",
         "the owner loses 300 Crystals per Mutagen flag on the field"),
        (r"\bAt end of owner's turn: foe loses 300 Crystals per Mutagen flag on field\b",
         "At end of the owner's turn, foe loses 300 Crystals per Mutagen flag on the field"),
        (r"\bAt end of owner's turn: if this card survived\b",
         "At end of the owner's turn, if this card survived"),
        (r"\bWhenever owner uses Tech, \+50 ATK&DEF until foe's turn ends\b",
         "Whenever the owner uses Tech, this card gains +50 ATK&DEF until foe's turn ends"),
        (r"\bWhenever foe activates Tech, \+10 ATK&DEF permanently\b",
         "Whenever foe activates Tech, this card gains +10 ATK&DEF permanently"),
        (r"\bIf this card defends, flip a coin\b",
         "If this card defends, flip a coin"),
        (r"\bPrevent a Miner or Mining card from being destroyed; destroy this card instead\b",
         "Prevent a Miner or Mining card from being destroyed; destroy this card instead"),
        (r"\bAfter attacking, this card cannot attack during their next turn\b",
         "After attacking, this card cannot attack on its next turn"),
        (r"\bAfter attacking, this card cannot attack during owner next turn\b",
         "After attacking, this card cannot attack on the owner's next turn"),
        (r"\bAfter attacking, cannot attack next owner turn\b",
         "After attacking, this card cannot attack on the owner's next turn"),
    ]
    for pat, rep in reps:
        t = re.sub(pat, rep, t, flags=re.I)
    t = re.sub(
        r"If the owner loses 500 or more Crystals, the owner recovers 300 Crystals",
        "If the owner loses 500 or more Crystals, the owner recovers 300 Crystals",
        t,
    )
    return t


def refine_text(text: str | None, *, keep_placeholders: bool = False, sheet: str = "") -> str:
    if text is None:
        return ""
    t = str(text).strip()
    if not t or t.lower() == "none":
        return t

    if keep_placeholders and "???" in t:
        t = re.sub(r"\bCrystal cost\b", "cost", t, flags=re.I)
        return finalize_ability_text(apply_shorthand_symbols(t))

    # Order matters
    reps: list[tuple[str, str]] = [
        (r"\bvs\.?\s*Chaos Affinity\b", "vs Chaos"),
        (r"\bvs\.?\s*Nature Affinity\b", "vs Nature"),
        (r"\bvs\.?\s*Arcane Affinity\b", "vs Arcane"),
        (r"\bvs\.?\s*Anima Affinity\b", "vs Anima"),
        (r"\bvs\.?\s*Bio Affinity\b", "vs Bio"),
        (r"\bvs\.?\s*Cosmic Affinity\b", "vs Cosmic"),
        (r"\bvs\.?\s*Divine Affinity\b", "vs Divine"),
        (r"\bvs\.?\s*Non-Arcane Affinity\b", "vs non-Arcane"),
        (r"\bvs\.?\s*Non-Anima Affinity\b", "vs non-Anima"),
        (r"\bface-up\b", "Exposed"),
        (r"\bface up\b", "Exposed"),
        (r"\bface-down defender\b", "face-down defender"),
        (r"\bface-down card\b", "face-down card"),
        (r"\bface-down unit\b", "face-down unit"),
        (r"\bface-down cell\b", "face-down cell"),
        (r"\bexposed card\b", "Exposed card"),
        (r"\bexposed cards\b", "Exposed cards"),
        (r"\bexposed unit\b", "Exposed unit"),
        (r"\bexposed units\b", "Exposed units"),
        (r"\bexposed ally units\b", "Exposed ally units"),
        (r"\bexposed ally unit\b", "Exposed ally unit"),
        (r"\bexposed foe\b", "Exposed foe"),
        (r"\bexposed ally\b", "Exposed ally"),
        (r"\bexposed cell\b", "Exposed cell"),
        (r"\bexposed cells\b", "Exposed cells"),
        (r"\bexposed position\b", "Exposed position"),
        (r"\bexposed turn\b", "Exposed turn"),
        (r"\bexposed card\b", "Exposed card"),
        (r"\bonce exposed\b", "Once Exposed"),
        (r"\bOnce exposed\b", "Once Exposed"),
        (r"\bwhile exposed\b", "while Exposed"),
        (r"\bWhile exposed\b", "While Exposed"),
        (r"\bwhen exposed\b", "when Exposed"),
        (r"\bWhen exposed\b", "When Exposed"),
        (r"\bif exposed\b", "if Exposed"),
        (r"\bIf exposed\b", "If Exposed"),
        (r"\bnot exposed\b", "not Exposed"),
        (r"\bget exposed\b", "become Exposed"),
        (r"\bturn exposed\b", "Expose"),
        (r"\bturn that card exposed\b", "Expose that card"),
        (r"\bsquare\b", "cell"),
        (r"\bsquares\b", "cells"),
        (r"\bFlip (\d+) coin\b", r"Flip \1 coins"),
        (r"\bper each head\(s\)\b", "per Head"),
        (r"\bfor each head\(s\)\b", "per Head"),
        (r"\bIf both are head\b", "If both are Heads"),
        (r"\bboth of them is head\b", "both are Heads"),
        (r"\bboth are head\b", "both are Heads"),
        (r"\bif both are head\b", "If both are Heads"),
        (r"\bAlways head\b", "Always Heads"),
        (r"\bIf head,\b", "Heads:"),
        (r"\bif head,\b", "Heads:"),
        (r"\bIf tail,\b", "Tails:"),
        (r"\bif tail,\b", "Tails:"),
        (r"\bIf head\b", "Heads:"),
        (r"\bIf tail\b", "Tails:"),
        (r"\bif head\b", "Heads:"),
        (r"\bif tail\b", "Tails:"),
        (r"\bHead:\s*", "Heads: "),
        (r"\bTail:\s*", "Tails: "),
        (r"\bTail -", "Tails: -"),
        (r"\bHeads:,\s*", "Heads: "),
        (r"\bTails:,\s*", "Tails: "),
        (r"\bboth player\b", "both players"),
        (r"\battacker have\b", "attacker has"),
        (r"\btrapper have\b", "trapper has"),
        (r"\bthe owner gain\b", "the owner gains"),
        (r"\bthe owner use\b", "the owner uses"),
        (r"\bthe owner have\b", "the owner has"),
        (r"\bfoe lose\b", "foe loses"),
        (r"\bfoe choose\b", "foe chooses"),
        (r"\bfoe must chooses\b", "foe must choose"),
        (r"\bfoe don't\b", "foe doesn't"),
        (r"\bAttacker end\b", "Attacker ends"),
        (r"\bThe trapper choose\b", "The trapper chooses"),
        (r"\bTrapper reveal\b", "The trapper reveals"),
        (r"\bTrapper switch\b", "The trapper swaps"),
        (r"\bTrapper select\b", "The trapper selects"),
        (r"\bAttacker choose\b", "The attacker chooses"),
        (r"\bThis player choose\b", "Choose"),
        (r"\bThis player select\b", "Select"),
        (r"\bThis player\. Choose\b", "Choose"),
        (r"\b on the this player", " on this player"),
        (r"\bThis bonus do not\b", "This bonus does not"),
        (r"\breveal (\d+) square\b", r"Reveal \1 cells"),
        (r"\b(\d+) square on\b", r"\1 cells on"),
        (r"\bfoe's cell\b", "foe's cells"),
        (r"\bReveal 1 foe's cells\b", "Reveal 1 foe cell"),
        (r"\bReveal 1 of foe's cells\b", "Reveal 1 foe cell"),
        (r"\bReveal 3 of foe's cells\b", "Reveal 3 foe cells"),
        (r"\b0-cost traps\b", "0-cost traps"),
        (r"\b0-cost trap\b", "0-cost trap"),
        (r"\b0-cost Traps\b", "0-cost traps"),
        (r"\bVenom Flag\b", "Venom flag"),
        (r"\bMutagen Flag\b", "Mutagen flag"),
        (r"\bprincess flag\b", "Princess flag"),
        (r"\bprinces flag\b", "Princess flag"),
        (r"\bacid flag\b", "Acid flag"),
        (r"\btech cards\b", "Tech cards"),
        (r"\btech card\b", "Tech card"),
        (r"\btech stack\b", "Tech Stack"),
        (r"\btech\b", "Tech"),
        (r"This card is unaffected by Tech cards\.", "Immune to Tech."),
        (r"This card is unaffected by Trap cards\.?", "Immune to traps."),
        (r"This card is unaffected by 0-cost traps\.?", "Immune to 0-cost traps."),
        (r"This card is not affected by 0-cost traps\.?", "Immune to 0-cost traps."),
        (r"This card is not affected by tech\.?", "Immune to Tech."),
        (r"This card is not affected by traps\.?", "Immune to traps."),
        (r"This Unit cannot be destroyed by Traps\.", "Immune to traps."),
        (r"This Unit cannot be destroyed by Tech Cards\.", "Immune to Tech."),
        (r"Immune to 0-cost Traps\.", "Immune to 0-cost traps."),
        (r"  +", " "),
    ]
    for pat, rep in reps:
        t = re.sub(pat, rep, t, flags=re.IGNORECASE if pat[0].islower() else 0)

    # cost phrasing: keep plural Crystals when standalone
    t = re.sub(r"\bpay no cost\b", "pay no cost", t, flags=re.I)
    t = re.sub(r"\(No paid cost\)\.?\s*", "without paying cost. ", t, flags=re.I)
    t = re.sub(r"\.\s*No paid cost\.?", " without paying cost.", t, flags=re.I)
    t = re.sub(r"\bNo paid cost\.?", "without paying cost.", t, flags=re.I)
    t = re.sub(r"\bno paid cost\.?", "without paying cost.", t, flags=re.I)
    t = re.sub(r"\bFoe pay no cost\b", "Foe pays no cost", t, flags=re.I)
    t = re.sub(r"\bfoe pay no cost\b", "Foe pays no cost", t, flags=re.I)
    t = re.sub(r"\bOnce only,\s*", "Once: ", t, flags=re.I)
    t = re.sub(r"\bOnce, ", "Once: ", t)
    t = re.sub(
        r"\brevive 1 unit to any (?:unoccupied or empty|empty) cell(?: in Exposed position| Exposed| in exposed position)?\.\s*",
        "revive 1 unit. ",
        t,
        flags=re.I,
    )
    t = re.sub(r"\bIn Reckoning, ", "In Reckoning: ", t)
    t = re.sub(r"\bUntil this turn ends, ", "This turn: ", t)
    t = re.sub(r"\bUntil the end of this turn, ", "This turn: ", t)
    t = re.sub(r"\bUntil the end of trapper's turn, ", "Until trapper's turn ends: ", t)
    t = re.sub(r"\bUntil the end of foe's turn, ", "Until foe's turn ends: ", t)
    t = re.sub(r"\bUntil foe's turn ends, ", "Until foe's turn ends: ", t)

    t = re.sub(r"\bHeads:,+", "Heads:", t)
    t = re.sub(r"\bTails:,+", "Tails:", t)
    t = re.sub(r"Heads:\s+,", "Heads:", t)
    t = re.sub(r"Tails:\s+,", "Tails:", t)

    # Never use You/Your — ambiguous for the opponent reading the card.
    t = re.sub(r"\byou may\b", "owner may", t, flags=re.I)
    t = re.sub(r"\byou can\b", "owner can", t, flags=re.I)
    t = re.sub(r"\byou must\b", "owner must", t, flags=re.I)
    t = re.sub(r"\byou already\b", "this player already", t, flags=re.I)
    t = re.sub(r"\byour\b", "this player's", t, flags=re.I)
    t = re.sub(r"\byou\b", "this player", t, flags=re.I)

    t = apply_grammar_shorten(t)
    t = re.sub(r"\bCrystal cost\b", "cost", t, flags=re.I)

    return finalize_ability_text(apply_shorthand_symbols(t.strip()))


def apply_manual(name: str, field: str, text: str, *, sheet: str = "") -> tuple[str, str]:
    entry = MANUAL.get(name, {})
    if field == "Partial Ability" and "Partial Ability" in entry:
        text_out = re.sub(r"\bCrystal cost\b", "cost", entry["Partial Ability"], flags=re.I)
        return finalize_ability_text(apply_shorthand_symbols(text_out)), entry.get("comment", "manual override")
    key = "Ability" if field in ("Ability", "Partial Ability") else "Full Ability"
    if key in entry or "Ability" in entry and field == "Ability":
        manual_key = key if key in entry else "Ability"
        if manual_key in entry:
            text_out = re.sub(r"\bCrystal cost\b", "cost", entry[manual_key], flags=re.I)
            return finalize_ability_text(apply_shorthand_symbols(text_out)), entry.get("comment", "manual override")
    new = refine_text(text, keep_placeholders=(field == "Partial Ability"), sheet=sheet)
    reasons = _reasons(text or "", new)
    comment = "; ".join(reasons) if reasons else ""
    return new, comment


def process_simple_sheet(ws, ability_col: int, old_ability_col: int, comment_col: int, *, sheet: str) -> None:
    ws.cell(row=2, column=old_ability_col, value="Old Ability")
    ws.cell(row=2, column=comment_col, value="Comment")
    for row in range(3, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name:
            continue
        old = ws.cell(row=row, column=ability_col).value
        if old is None or str(old).strip() == "":
            continue
        old_s = str(old).strip()
        if old_s.lower() == "none":
            continue
        ws.cell(row=row, column=old_ability_col, value=old_s)
        new, comment = apply_manual(str(name).strip(), "Ability", old_s, sheet=sheet)
        if new != old_s:
            ws.cell(row=row, column=ability_col, value=new)
            ws.cell(row=row, column=comment_col, value=comment)


def process_union_sheet(
    ws,
    partial_col: int,
    full_col: int,
    old_partial_col: int,
    old_full_col: int,
    comment_col: int,
) -> None:
    ws.cell(row=2, column=old_partial_col, value="Old Partial Ability")
    ws.cell(row=2, column=old_full_col, value="Old Full Ability")
    ws.cell(row=2, column=comment_col, value="Comment")
    for row in range(3, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name:
            continue
        name_s = str(name).strip()
        comments: list[str] = []

        partial_old = ws.cell(row=row, column=partial_col).value
        if partial_old and str(partial_old).strip().lower() not in ("none", ""):
            partial_s = str(partial_old).strip()
            ws.cell(row=row, column=old_partial_col, value=partial_s)
            partial_new, c = apply_manual(name_s, "Partial Ability", partial_s, sheet="Union")
            if partial_new != partial_s:
                ws.cell(row=row, column=partial_col, value=partial_new)
                comments.append("Partial: " + (c or "; ".join(_reasons(partial_s, partial_new))))

        full_old = ws.cell(row=row, column=full_col).value
        if full_old and str(full_old).strip().lower() not in ("none", ""):
            full_s = str(full_old).strip()
            ws.cell(row=row, column=old_full_col, value=full_s)
            full_new, c = apply_manual(name_s, "Full Ability", full_s, sheet="Union")
            if full_new != full_s:
                ws.cell(row=row, column=full_col, value=full_new)
                comments.append("Full: " + c)

        if comments:
            ws.cell(row=row, column=comment_col, value=" | ".join(comments))


ABILITY_TEXT_RULES: list[tuple[str, str, str, str]] = [
    ("Priority", "Rule", "Use", "Avoid / Notes"),
    ("", "Brevity", "As short as possible", "Correct English; clear to both players"),
    ("", "", "", ""),
    ("Terminology", "Rule", "Use", "Avoid / Notes"),
    ("", "Visible state", "Exposed", "face-up; lowercase exposed (same as code face_up)"),
    ("", "Hidden state", "face-down", "Keep lowercase"),
    ("", "Grid", "cell", "square, slot"),
    ("", "Characters", "Preserve card, unit, trap from source; monster→unit", "Do not cross-rename card↔unit"),
    ("", "Battle phase", "Reckoning", "in battle, when battling"),
    ("", "Trap owner", "Trapper", "Trapper = trap owner"),
    ("", "Perspective", "owner, this player, foe, attacker", "You, Your"),
    ("", "Coin result", "Heads: / Tails:", "if head, Head:, head(s)"),
    ("", "Coin count", "Flip 3 coins", "Flip 3 coin"),
    ("", "Currency amount", "500 Crystals", "Crystal as standalone amount"),
    ("", "Paying cost", "500 cost; no cost", "Use cost not Crystal cost; Crystals when amount only"),
    ("", "Flags", "Venom flag, Mutagen flag, Princess flag", "Venom Flag, ALL CAPS"),
    ("", "Free traps", "0-cost traps", "0 cost traps"),
    ("", "Tech type", "Tech", "lowercase tech"),
    ("", "Immunity", "Immune to Tech. / Immune to traps.", "Long unaffected-by wording"),
    ("", "", "", ""),
    ("Stats & shorthand", "Rule", "Use", "Avoid / Notes"),
    ("", "Combined stats", "ATK&DEF", "ATK and DEF, ATK/DEF"),
    ("", "Comparisons", "≥ ≤ for cost thresholds; at least 1 for counts", "≥1 for counts; = # for counting"),
    ("", "Examples", "cost ≥800; at least 1 Exposed card; ATK&DEF becomes 0", "ATK&DEF=0; = # Heads; half symbol"),
    ("", "Becomes", "ATK&DEF becomes 0", "Do not shorthand becomes to ="),
    ("", "Half", "half of / halve", "No ½ symbol"),
    ("", "Colons", "Single : only", "No ::"),
    ("", "Grammar", "Short + correct English", "Missing articles/subjects; bare fragments"),
    ("", "Examples", "After a successful attack; At turn start; on the field", "After successful attack; Turn start"),
    ("", "Counting", "for each Head; at least 1", "= # Heads"),
    ("", "Revive", "Revive Pyromancer", "No 'from void' in text"),
    ("", "Union trigger", "Summoned:", "Once Union, Upon union, On Union:"),
    ("", "Matchups", "+50 ATK vs Chaos", "vs Chaos Affinity"),
    ("", "Agreement", "both players; foe chooses; attacker has; owner may", "both player; attacker have"),
    ("", "Coin payout", "per Head", "per each head(s)"),
    ("", "", "", ""),
    ("Structure", "Rule", "Use", "Avoid / Notes"),
    ("", "One-shot", "Once:", "Once, only"),
    ("", "Battle", "In Reckoning:", "In Reckoning, only"),
    ("", "Duration", "This turn: / Until foe's turn ends:", "Overlong duration phrasing"),
    ("", "End trigger", "End of turn:", "At the end of this turn, when shorter works"),
    ("", "", "", ""),
    ("Game rules in text", "Rule", "Use", "Avoid / Notes"),
    ("", "Flag on face-down", "Flag applies; unit becomes Exposed permanently", "Temporary peek"),
    ("", "Plant-29", "Start of owner's turn: Flip a coin. Head: put Venom Flag on 1 exposed ally or foe card. Tail: put Mutagen Flag on any of your unit.", "Synced to card_data.xlsx"),
    ("", "Death Cobra", "End of turn: Choose 1 foe card. Put Venom flag on it.", "Works on face-down targets"),
    ("", "", "", ""),
    ("Union", "Rule", "Use", "Avoid / Notes"),
    ("", "Partial Ability", "Keep ??? teasers in UI", ""),
    ("", "Full Ability", "One ability per card unless design says otherwise", "Do not merge unrelated effects"),
    ("", "", "", ""),
    ("Avoid", "Rule", "Use", "Avoid / Notes"),
    ("", "", "", "You / Your"),
    ("", "", "", "Ambiguous their"),
    ("", "", "", "Broken grammar: the this player's"),
    ("", "", "", "Mixing or cross-renaming card/unit/trap"),
    ("", "", "", ""),
    ("Refinement", "Rule", "Use", "Avoid / Notes"),
    ("", "Source", "context/card_data.xlsx", "Master card data spreadsheet"),
    ("", "Output cols", "Ability, Old Ability, Comment", "Union: Old Partial/Full Ability too"),
    ("", "", "", ""),
    ("Confirmed design", "Rule", "Use", "Avoid / Notes"),
    ("", "Exposed vs face-up", "Exposed on card text", "Same as code face_up"),
    ("", "Revive", "Revive (void implied)", "No 'from void' in text"),
    ("", "Affinity immunity", "In Reckoning: not destroyed by X cards", "Not global Immune to (traps/Tech)"),
    ("", "Burning Phoenix", "Non-Union card immunity in Reckoning only", "Traps/Tech not included; Tech destroys"),
    ("", "Electrogazer", "Disable 0-cost traps", "Not destroy"),
    ("", "Dreamcatcher", "Triggers when foe uses Tech", "Cancel that Tech immediately"),
    ("", "Damage", "500 Crystals", "damage = Crystals on card text"),
    ("", "Silver Dragon", "1000 cost per attack", "Pay each attack"),
    ("", "Night Whisperer", "Name contains wisp", "Archetype by name substring"),
    ("", "Archbishop", "Sacrifice Divine card, even if face-down", ""),
    ("", "Lab Bloater", "Both players pay no cost", ""),
    ("", "Venom Toad", "Destroy Venom-flag units every Reckoning", ""),
]


def add_ability_text_rules_sheet(wb: openpyxl.Workbook) -> None:
    title = "Ability Text Rules"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title, 0)
    for r, row in enumerate(ABILITY_TEXT_RULES, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 48


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"Missing source: {SRC}")
    shutil.copy2(SRC, OUT)
    wb = openpyxl.load_workbook(OUT)

    for sheet in ("Unit", "Trap", "Tech"):
        ws = wb[sheet]
        ws.insert_cols(9, amount=2)  # Old Ability, Comment (after Ability col 8)
        process_simple_sheet(ws, ability_col=8, old_ability_col=9, comment_col=10, sheet=sheet)

    ws_u = wb["Union"]
    ws_u.insert_cols(10, amount=3)  # Old Partial, Old Full, Comment
    process_union_sheet(
        ws_u,
        partial_col=8,
        full_col=9,
        old_partial_col=10,
        old_full_col=11,
        comment_col=12,
    )

    add_ability_text_rules_sheet(wb)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
