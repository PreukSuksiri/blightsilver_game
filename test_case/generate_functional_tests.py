#!/usr/bin/env python3
"""
Generate code-accurate functional test cases for Demo=Yes cards.
Sources: CardDatabase.gd, UnionDatabase.gd, demo_flags.json, BattleResolver.gd, TurnManager.gd
"""

from __future__ import annotations

import json
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
OUT = Path(__file__).resolve().parent / "functional"


# ── Parsers ──────────────────────────────────────────────────

def load_demo_names() -> set[str]:
    return set(json.loads((ROOT / "data/demo_flags.json").read_text()).keys())


def load_union_demo_from_excel() -> dict[str, dict]:
    """Excel Demo=Yes unions — includes cards not yet in UnionDatabase.gd."""
    import openpyxl
    wb = openpyxl.load_workbook(ROOT / "context/card_data_demo.xlsx", data_only=True)
    ws = wb["Union"] if "Union" in wb.sheetnames else wb["Character"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else "" for h in rows[1]]
    idx = {h: i for i, h in enumerate(headers) if h}
    out: dict[str, dict] = {}
    for row in rows[2:]:
        if not row or not row[0]:
            continue
        if not row[idx["Demo"]] or str(row[idx["Demo"]]).strip().lower() != "yes":
            continue
        name = str(row[0]).strip()
        out[name] = {
            "card_type": "Union",
            "name": name,
            "affinity": str(row[idx.get("Affinity", 5)] or "ANIMA"),
            "atk": int(row[idx["ATK"]] or 0),
            "def": int(row[idx["DEF"]] or 0),
            "summon_cost": None,
            "ability": str(row[idx.get("Full Ability", 8)] or row[idx.get("Partial Ability", 7)] or "None"),
            "params_raw": "{}",
            "description": str(row[idx.get("Full Ability", 8)] or row[idx.get("Partial Ability", 7)] or ""),
            "formula": str(row[idx.get("Full Formula", 10)] or row[idx.get("Partial Formula", 9)] or ""),
            "in_union_database": False,
        }
    return out


def merge_union_cards(unions_db: dict[str, dict], excel: dict[str, dict]) -> list[dict]:
    merged: dict[str, dict] = {}
    for name, ex in excel.items():
        if name in unions_db:
            u = {**unions_db[name], "in_union_database": True, "formula": ex.get("formula", "")}
            merged[name] = u
        else:
            merged[name] = {
                **ex,
                "ability": "EXCEL_ONLY_NOT_IN_UNION_DATABASE",
                "description": ex.get("description") or "See Excel; not yet in UnionDatabase.gd",
            }
    return list(merged.values())


def parse_card_database() -> dict[str, dict]:
    text = (ROOT / "autoload/CardDatabase.gd").read_text()
    cards: dict[str, dict] = {}

    def load_section(func: str, card_type: str):
        idx = text.find(f"func {func}")
        s = text.find("var defs: Array = [", idx)
        e = text.find("\n\t]", s)
        for part in re.split(r"\n\t\t\[", text[s : e + 4])[1:]:
            part = "[" + part
            name_m = re.match(r'\["([^"]+)"', part)
            if not name_m:
                continue
            name = name_m.group(1)
            if card_type == "character":
                m = re.match(
                    r'\["([^"]+)",\s*CharacterData\.Affinity\.(\w+),\s*(\d+),\s*(\d+),\s*(\d+),\s*'
                    r"CharacterData\.AbilityType\.(\w+),\s*(\{.*?\}),\s*\n\t\t\t\"([^\"]*)\"",
                    part,
                    re.S,
                )
                if m:
                    cards[name] = {
                        "card_type": "Character",
                        "name": name,
                        "affinity": m.group(2),
                        "atk": int(m.group(3)),
                        "def": int(m.group(4)),
                        "cost": int(m.group(5)),
                        "ability": m.group(6),
                        "params_raw": m.group(7),
                        "description": m.group(8),
                    }
            elif card_type == "trap":
                m = re.match(
                    r'\["([^"]+)",\s*(\d+),\s*TrapData\.TrapEffectType\.(\w+),\s*(\{.*?\}),\s*\n\t\t\t"([^"]*)"',
                    part,
                    re.S,
                )
                if m:
                    cards[name] = {
                        "card_type": "Trap",
                        "name": name,
                        "cost": int(m.group(2)),
                        "effect": m.group(3),
                        "params_raw": m.group(4),
                        "description": m.group(5),
                    }
            elif card_type == "tech":
                m = re.match(
                    r'\["([^"]+)",\s*(\d+),\s*TechCardData\.TechEffectType\.(\w+),\s*(\{.*?\}),\s*'
                    r'(?:"([^"]*)",\s*)?\n\t\t\t"([^"]*)"',
                    part,
                    re.S,
                )
                if m:
                    cards[name] = {
                        "card_type": "Tech",
                        "name": name,
                        "cost": int(m.group(2)),
                        "effect": m.group(3),
                        "params_raw": m.group(4),
                        "required_prior": m.group(5) or "",
                        "description": m.group(6),
                    }

    load_section("_load_characters", "character")
    load_section("_load_traps", "trap")
    load_section("_load_tech_cards", "tech")
    return cards


def parse_union_database() -> dict[str, dict]:
    text = (ROOT / "autoload/UnionDatabase.gd").read_text()
    unions: dict[str, dict] = {}
    pattern = re.compile(
        r'_add\("([^"]+)",\s*A\.(\w+),\s*(\d+),\s*(\d+),\s*(\d+),\s*R\.\w+,\s*'
        r"AB\.(\w+),\s*(\{.*?\}),\s*\"([^\"]*)\"",
        re.S,
    )
    for m in pattern.finditer(text):
        unions[m.group(1)] = {
            "card_type": "Union",
            "name": m.group(1),
            "affinity": m.group(2),
            "atk": int(m.group(3)),
            "def": int(m.group(4)),
            "summon_cost": int(m.group(5)),
            "ability": m.group(6),
            "params_raw": m.group(7),
            "description": m.group(8),
        }
    return unions


def slug(name: str) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "-", name.strip()).strip("-")
    return s or "Unknown"


def parse_params(raw: str) -> dict:
    """Best-effort parse of GDScript dict literal to Python-ish values."""
    if not raw or raw == "{}":
        return {}
    out: dict = {}
    for m in re.finditer(r'"(\w+)":\s*([^,}\]]+)', raw):
        key, val = m.group(1), m.group(2).strip()
        if val.startswith('"'):
            out[key] = val.strip('"')
        elif val.startswith("CharacterData.Affinity."):
            out[key] = val.split(".")[-1]
        elif val.startswith("["):
            affs = re.findall(r"CharacterData\.Affinity\.(\w+)", val)
            out[key] = affs
        elif val in ("true", "false"):
            out[key] = val == "true"
        else:
            try:
                out[key] = int(val)
            except ValueError:
                out[key] = val
    return out


def fmt_case(
    case_id: str,
    desc: str,
    impl: list[str],
    pre: list[str],
    steps: list[str],
    expected: list[str],
    verify: list[str] | None = None,
) -> str:
    lines = [
        f"Test Case ID: {case_id}",
        "Description:",
        desc,
        "Implementation Reference:",
    ]
    for i in impl:
        lines.append(f"- {i}")
    lines.append("Preconditions:")
    for p in pre:
        lines.append(f"- {p}")
    lines.append("Steps:")
    for i, s in enumerate(steps, 1):
        lines.append(f"Step {i}: {s}")
    lines.append("Expected Result:")
    for e in expected:
        lines.append(f"- {e}")
    if verify:
        lines.append("Verification (automated):")
        for v in verify:
            lines.append(f"- {v}")
    return "\n".join(lines)


BASE_PRE = [
    "Godot battle_test or Daily Dungeon; `CardDatabase` loaded.",
    "Both players STARTING_CRYSTALS=5000 unless test specifies otherwise.",
    "Disable `bare_hands_brawling` dungeon modifier (cancels character abilities in BattleResolver).",
]


def card_header(card: dict) -> str:
    ct = card["card_type"]
    lines = [f"Card Name: {card['name']}", f"Type: {ct}"]
    if ct == "Character":
        lines += [
            f"Stats: ATK={card['atk']} DEF={card['def']} Cost={card['cost']} Affinity={card['affinity']}",
            f"AbilityType: {card['ability']}",
            f"ability_params: {parse_params(card.get('params_raw', ''))}",
            f"Description: {card.get('description', '')}",
        ]
    elif ct == "Trap":
        lines += [
            f"Trap Cost: {card['cost']}",
            f"TrapEffectType: {card['effect']}",
            f"effect_params: {parse_params(card.get('params_raw', ''))}",
            f"Description: {card.get('description', '')}",
        ]
    elif ct == "Tech":
        lines += [
            f"Tech Cost: {card['cost']}",
            f"TechEffectType: {card['effect']}",
            f"effect_params: {parse_params(card.get('params_raw', ''))}",
            f"Description: {card.get('description', '')}",
        ]
    elif ct == "Union":
        lines += [
            f"Stats: ATK={card['atk']} DEF={card['def']} summon_cost={card['summon_cost']} Affinity={card['affinity']}",
            f"AbilityType: {card['ability']}",
            f"ability_params: {parse_params(card.get('params_raw', ''))}",
            f"Description: {card.get('description', '')}",
        ]
    lines.append("Test Cases:")
    return "\n".join(lines)


# ── Ability-specific functional test builders ─────────────────

def build_character_tests(card: dict) -> list[str]:
    name = card["name"]
    sid = slug(name)
    ab = card["ability"]
    p = parse_params(card.get("params_raw", ""))
    atk, df, cost = card["atk"], card["def"], card["cost"]
    cases: list[str] = []
    n = 1

    def add(desc, impl, pre, steps, expected, verify=None):
        nonlocal n
        cid = f"TC-FUNC-{sid}-{n:03d}"
        n += 1
        pre_all = BASE_PRE + pre
        cases.append(fmt_case(cid, desc, impl, pre_all, steps, expected, verify))

    # ── BattleResolver stat modifiers ──
    if ab == "ATK_BONUS_VS_AFFINITY":
        bonus = p.get("bonus", 0)
        aff = p.get("affinity", "?")
        add(
            f"{name}: effective ATK +{bonus} vs {aff} defender",
            [
                "BattleResolver._get_effective_atk()",
                f"CharacterData.AbilityType.{ab}",
                f"ability_params: affinity={aff}, bonus={bonus}",
            ],
            [
                f"Attacker {name} (base ATK {atk}) face-up.",
                f"Defender: any {aff} character with DEF < {atk + bonus}.",
            ],
            [
                f"Call BattleResolver.resolve_battle({name}, {aff} defender, dice_roll=3, ...).",
                "Inspect battle overlay / result.attacker_atk_used.",
            ],
            [
                f"result.attacker_atk_used == {atk + bonus} (base {atk} + bonus {bonus}).",
                f"result.attacker_atk_delta == {bonus}.",
                "Defender destroyed when effective ATK > DEF.",
            ],
            [
                f"assert_eq(result.attacker_atk_used, {atk + bonus})",
                f"assert_true(result.defender_destroyed)",
            ],
        )
        add(
            f"{name}: no bonus vs non-{aff} defender",
            ["BattleResolver._get_effective_atk()", f"AbilityType.{ab}"],
            [f"Defender affinity != {aff} (e.g. ANIMA Wandering Swordsman)."],
            [f"Attack non-{aff} target with {name}."],
            [
                f"result.attacker_atk_used == {atk} (no affinity bonus).",
                f"result.attacker_atk_delta == 0.",
            ],
            [f"assert_eq(result.attacker_atk_used, {atk})"],
        )

    elif ab == "ATK_DEF_BONUS_VS_AFFINITY":
        ba, bd = p.get("atk", 0), p.get("def", 0)
        aff = p.get("affinity", "?")
        add(
            f"{name}: +{ba} ATK / +{bd} DEF vs {aff}",
            ["BattleResolver._get_effective_atk/_get_effective_def", f"AbilityType.{ab}"],
            [f"Battle {name} (ATK {atk}) vs {aff} defender; or {name} defends vs {aff} attacker."],
            ["Resolve battle; read effective stats."],
            [
                f"Attacking: attacker_atk_used == {atk + ba}.",
                f"Defending: defender_def_used == {df + bd} when {name} is defender.",
            ],
            [],
        )

    elif ab == "ATK_DEF_BONUS_VS_NON_AFFINITY":
        bonus = p.get("atk", p.get("bonus", 0))
        own_aff = p.get("affinity", card["affinity"])
        add(
            f"{name}: +{bonus} ATK/DEF vs non-{own_aff}",
            ["BattleResolver._get_effective_atk/_get_effective_def", f"AbilityType.{ab}"],
            [f"Opponent character affinity != {own_aff}."],
            [f"Attack or defend with {name}."],
            [f"When opponent affinity != {own_aff}: +{bonus} to ATK (attack) or DEF (defend)."],
            [],
        )

    elif ab == "ATK_BONUS_IF_AFFINITY_ON_FIELD":
        bonus = p.get("bonus", 0)
        aff = p.get("affinity", "?")
        add(
            f"{name}: +{bonus} ATK when face-up {aff} ally on field",
            ["BattleResolver._get_effective_atk() scans GameState.grids", f"AbilityType.{ab}"],
            [
                f"Place another face-up {aff} character on attacker's grid.",
                f"{name} attacks (does not need to target {aff}).",
            ],
            ["Run calculate_field_bonuses if needed; resolve battle."],
            [f"attacker_atk_used == {atk + bonus}."],
            [],
        )

    elif ab == "DEF_BONUS_IF_AFFINITY_ON_FIELD":
        bonus = p.get("def", p.get("bonus", 0))
        aff = p.get("affinity", "?")
        add(
            f"{name}: +{bonus} DEF when face-up {aff} on own field",
            ["BattleResolver._get_effective_def()", f"AbilityType.{ab}"],
            [f"Face-up {aff} ally on field; opponent attacks {name}."],
            ["Resolve defense."],
            [f"defender_def_used == {df + bonus}."],
            [],
        )

    elif ab == "ATK_DEF_BONUS_IF_UNION_ON_FIELD":
        ba, bd = p.get("atk", 0), p.get("def", 0)
        add(
            f"{name}: +{ba}/+{bd} when Union on own field",
            ["BattleResolver._get_effective_atk() checks is_union face-up ally", f"AbilityType.{ab}"],
            ["Summon or place face-up Union on attacker's field."],
            [f"Attack with {name}."],
            [f"attacker_atk_used == {atk + ba}; effective DEF bonus {bd} when defending."],
            [],
        )

    elif ab == "ATK_BONUS_VS_FACEDOWN":
        bonus = p.get("bonus", 0)
        add(
            f"{name}: +{bonus} ATK vs face-down defender",
            ["BattleResolver._get_effective_atk() checks not defender.face_up", f"AbilityType.{ab}"],
            ["Defender remains face-down at calculation time (Skeleton Archer scenario)."],
            [f"Attack face-down cell with {name}."],
            [f"If defender.face_up==false at calc: attacker_atk_used == {atk + bonus}."],
            [],
        )

    elif ab == "ATK_BOOST_VS_REVEALED":
        bonus = p.get("bonus", 0)
        add(
            f"{name}: +{bonus} ATK vs exposed defender",
            ["BattleResolver._get_effective_atk() checks defender.face_up", f"AbilityType.{ab}"],
            ["Defender was face-up before attack (defender_was_exposed=true)."],
            [f"Attack revealed defender with {name}."],
            [f"attacker_atk_used == {atk + bonus}."],
            [],
        )

    elif ab == "ATK_BONUS_VS_UNION":
        bonus = p.get("bonus", 0)
        add(
            f"{name}: +{bonus} ATK vs Union",
            ["BattleResolver._get_effective_atk() checks defender.is_union", f"AbilityType.{ab}"],
            ["Opponent has face-up Union (e.g. Gryphon Rider)."],
            [f"Attack Union with {name}."],
            [f"attacker_atk_used == {atk + bonus}."],
            [],
        )

    elif ab == "ATK_BONUS_VS_VENOM":
        bonus = p.get("bonus", 0)
        add(
            f"{name}: +{bonus} ATK vs venom-flagged target",
            ["BattleResolver._get_effective_atk() checks 'venom' in defender.flags", f"AbilityType.{ab}"],
            ['Set defender.flags = ["venom"] (via Death Cobra end-of-turn).'],
            [f"Attack flagged defender with {name}."],
            [f"attacker_atk_used == {atk + bonus}."],
            [],
        )

    elif ab == "ATK_BONUS_VS_CENTER_ZONE":
        zb, cb = p.get("zone_bonus", 20), p.get("center_bonus", 40)
        add(
            f"{name}: center zone ATK bonuses (+{zb} 3×3, +{cb} at 2,2)",
            ["BattleResolver._get_effective_atk() uses target_pos", f"AbilityType.{ab}"],
            [f"{name} on field; attack opponent cells at edge, center-zone, and (2,2)."],
            ["Pass target_pos to resolve_battle; compare attacker_atk_used each time."],
            [
                f"Edge target: attacker_atk_used == {atk}.",
                f"Center zone (not 2,2): attacker_atk_used == {atk + zb}.",
                f"Cell (2,2): attacker_atk_used == {atk + zb + cb}.",
            ],
            [],
        )

    elif ab == "ATK_PENALTY_WHEN_EXPOSED":
        pen = p.get("penalty", p.get("amount", 0))
        add(
            f"{name}: -{pen} ATK while face-up",
            ["BattleResolver._get_effective_atk() when attacker.face_up", f"AbilityType.{ab}"],
            [f"{name} face-up on field."],
            ["Attack any target."],
            [f"attacker_atk_used == max(0, {atk} - {pen}) == {max(0, atk - pen)}."],
            [],
        )

    elif ab == "DEF_ZERO_WHEN_EXPOSED":
        add(
            f"{name}: DEF=0 while face-up",
            ["BattleResolver._get_effective_def() returns 0 when defender.face_up", f"AbilityType.{ab}"],
            [f"{name} face-up; opponent attacks."],
            ["Resolve battle."],
            [f"defender_def_used == 0 regardless of base DEF {df}."],
            ["assert_eq(result.defender_def_used, 0)"],
        )

    elif ab == "ATTACKER_ATK_DEBUFF":
        amt = p.get("amount", 0)
        add(
            f"{name}: debuff attacker ATK by {amt} during battle",
            ["BattleResolver._get_effective_atk() defender branch", f"AbilityType.{ab}"],
            [f"{name} defends; attacker ATK > {df}."],
            ["Resolve battle."],
            [
                f"Attacker effective ATK reduced by {amt} (unless effect_nullified_until active).",
                "GameState.post_message contains 'Attacker loses N ATK'.",
            ],
            [],
        )

    elif ab == "ATTACK_STANCE_BOOST":
        b = p.get("atk_bonus", p.get("atk", 0))
        add(
            f"{name}: +{b} ATK in attack stance",
            ["BattleResolver._get_effective_atk()", f"AbilityType.{ab}"],
            [f"{name} attacks."],
            ["Resolve battle."],
            [f"attacker_atk_used == {atk + b}."],
            [],
        )

    elif ab == "DEFENSE_STANCE_BOOST":
        b = p.get("def_bonus", p.get("def", 0))
        add(
            f"{name}: +{b} DEF in defense stance",
            ["BattleResolver._get_effective_def()", f"AbilityType.{ab}"],
            [f"Opponent attacks {name}."],
            ["Resolve battle."],
            [f"defender_def_used == {df + b}."],
            [],
        )

    elif ab == "SWAP_ATK_DEF_WHEN_ATTACKING":
        add(
            f"{name}: uses DEF as ATK when attacking",
            ["BattleResolver._get_effective_atk() sets atk = get_effective_def()", f"AbilityType.{ab}"],
            [f"{name} ATK={atk} DEF={df} attacks."],
            ["Resolve battle."],
            [f"attacker_atk_used == effective DEF ({df} + temp/perm bonuses), not base ATK {atk}."],
            [],
        )

    elif ab == "ONE_USE_ATK_BOOST":
        b = p.get("bonus", 0)
        add(
            f"{name}: one-time +{b} ATK on first attack",
            [
                "BattleResolver._get_effective_atk() if not one_use_atk_boost_used",
                "TurnManager._apply_post_battle_effects sets one_use_atk_boost_used=true",
                f"AbilityType.{ab}",
            ],
            [f"{name} has one_use_atk_boost_used=false."],
            ["First attack: verify bonus; second attack: no bonus."],
            [
                f"First attack: attacker_atk_used == {atk + b}; one_use_atk_boost_used becomes true.",
                f"Second attack: attacker_atk_used == {atk}.",
            ],
            [],
        )

    elif ab == "ONE_USE_DEF_BOOST":
        b = p.get("bonus", 0)
        add(
            f"{name}: one-time +{b} DEF when defending",
            [
                "BattleResolver._get_effective_def()",
                "TurnManager marks one_use_def_boost_used after battle",
                f"AbilityType.{ab}",
            ],
            [f"First defense with {name}."],
            ["Opponent attacks twice across two turns."],
            [
                f"First defense: defender_def_used == {df + b}.",
                f"Second defense: defender_def_used == {df}.",
            ],
            [],
        )

    elif ab == "ONE_USE_TEMP_BOOST_ATTACK_AND_DEFEND":
        ba, bd = p.get("atk_bonus", p.get("atk", 0)), p.get("def_bonus", p.get("def", 0))
        add(
            f"{name}: once +{ba} ATK attack / once +{bd} DEF defend",
            ["BattleResolver separate one_use flags for atk/def uses", f"AbilityType.{ab}"],
            [f"{name} unused one-use flags."],
            ["Attack once, then defend once on later turn."],
            [
                f"First attack includes +{ba} ATK.",
                f"First defend includes +{bd} DEF.",
                "Subsequent uses: no corresponding bonus.",
            ],
            [],
        )

    elif ab == "BOOST_PER_TYPED_CARD_ON_FIELD":
        abonus = p.get("atk_bonus", 0)
        dbonus = p.get("def_bonus", 0)
        filt = p.get("card_name_contains") or p.get("affinity", "")
        add(
            f"{name}: field passive +{abonus} ATK / +{dbonus} DEF per matching card",
            [
                "BattleResolver.calculate_field_bonuses() → _apply_field_ability_bonus",
                "Sets perm_atk_bonus / perm_def_bonus on source card",
                f"AbilityType.{ab}",
            ],
            [
                f"Place 2+ matching face-up allies (filter: {filt}) on attacker's field.",
                "Call calculate_field_bonuses(player) before battle.",
            ],
            [f"Inspect {name}.perm_atk_bonus and perm_def_bonus."],
            [
                f"perm_atk_bonus == {abonus} × matching_count.",
                f"perm_def_bonus == {dbonus} × matching_count.",
                "get_effective_atk/def includes perm bonuses.",
            ],
            [],
        )

    elif ab == "BOOST_PER_ANIMA_ON_FIELD":
        abonus = p.get("atk_bonus", 0)
        dbonus = p.get("def_bonus", 0)
        add(
            f"{name}: +{abonus}/+{dbonus} per other face-up Anima on field",
            ["BattleResolver._count_anima_cards()", f"AbilityType.{ab}"],
            ["2 other face-up Anima characters on field."],
            ["calculate_field_bonuses then battle."],
            [f"With 2 anima allies: perm_atk_bonus=={abonus*2}, perm_def_bonus=={dbonus*2}."],
            [],
        )

    elif ab == "DESTROY_IF_OPPONENT_AFFINITY":
        aff = p.get("affinity", "?")
        add(
            f"{name}: auto-destroy {aff} defender before ATK/DEF compare",
            ["BattleResolver._resolve_character_vs_character pre-compare branch", f"AbilityType.{ab}"],
            [f"Defender is {aff} character."],
            [f"{name} attacks {aff} defender."],
            [
                "result.defender_destroyed == true regardless of stats.",
                f"result.defender_crystal_loss == defender.crystal_cost.",
                "Normal ATK/DEF comparison skipped.",
            ],
            ["assert_true(result.defender_destroyed)"],
        )

    elif ab == "DESTROYED_IF_BATTLES_DIVINE":
        add(
            f"{name}: destroyed when battling Divine; halve stats after attack",
            [
                "BattleResolver pre-compare: attacker destroyed if defender.affinity==DIVINE",
                "_apply_post_attack_effects: halve_stats() via ability_type match",
                f"AbilityType.{ab}",
            ],
            ["Divine defender (e.g. Angel Gatekeeper) face-up."],
            [f"{name} attacks Divine character."],
            [
                "result.attacker_destroyed == true.",
                f"result.attacker_crystal_loss == {cost}.",
                "After any successful attack: current_atk and current_def halved permanently.",
            ],
            [],
        )

    elif ab == "DESTROY_SELF_VS_DIVINE_BOTH":
        add(
            f"{name}: self-destruct when battling Divine (attacker or defender role)",
            ["BattleResolver early-return branches for Divine match", f"AbilityType.{ab}"],
            ["Divine character involved in battle."],
            [f"Battle {name} vs Divine."],
            [f"{name} result.{ 'attacker' if 'Vampire' in name or 'Feral' in name else 'either' }_destroyed == true."],
            [],
        )

    elif ab == "COIN_FLIP_2_DESTROY_NON_AFFINITY":
        aff = p.get("affinity", card["affinity"])
        add(
            f"{name}: 2 coin flips; both heads destroys non-{aff} defender",
            ["BattleResolver uses randf()>=0.5 twice; skips normal battle on success", f"AbilityType.{ab}"],
            [f"Non-{aff} defender."],
            ["Run battle; log coin outcomes."],
            [
                "Both heads: defender_destroyed=true, no ATK/DEF comparison.",
                "Otherwise: normal battle proceeds.",
            ],
            ["Non-deterministic — run ≥20 iterations; verify both branches occur."],
        )

    elif ab == "COIN_FLIP_ATK_BOOST":
        b = p.get("bonus", 0)
        add(
            f"{name}: coin flip +{b} ATK this battle",
            ["BattleResolver._get_effective_atk() randf()>=0.5", f"AbilityType.{ab}"],
            [f"{name} attacks."],
            ["Resolve battle multiple times."],
            [f"Heads: attacker_atk_used == {atk + b}; Tails: attacker_atk_used == {atk}."],
            [],
        )

    elif ab == "COIN_FLIP_CANCEL_ATTACK":
        add(
            f"{name}: coin flip may cancel own attack",
            ["TurnManager pre-attack check AbilityType.COIN_FLIP_CANCEL_ATTACK", f"AbilityType.{ab}"],
            [f"{name} selected as attacker."],
            ["Initiate attack; observe pre-battle coin."],
            ["Tails: attack cancelled before battle; Heads: proceeds."],
            [],
        )

    elif ab == "COIN_FLIP_EXTRA_ATTACK":
        add(
            f"{name}: coin flip extra attack after battle",
            ["TurnManager._apply_post_battle_effects randi()%2", f"AbilityType.{ab}"],
            [f"{name} completes an attack."],
            ["Check attacks_remaining after battle."],
            ["Heads: attacks_remaining incremented by 1; Tails: no extra."],
            [],
        )

    elif ab == "COIN_FLIP_SWAP_POSITION":
        add(
            f"{name}: coin flip swap position after battle",
            ["BattleResolver sets pending_coin_flip_swap_position; TurnManager prompts swap", f"AbilityType.{ab}"],
            [f"{name} on field with another own character."],
            ["Complete battle; on heads choose swap target."],
            ["Grid positions exchange; cards retain stats."],
            [],
        )

    elif ab == "MUTAGEN_DESTROY_ATTACKER":
        add(
            f"{name}: with has_mutagen_flag, destroy attacker after surviving",
            [
                "BattleResolver post-compare: if not defender_destroyed and has_mutagen_flag",
                "Sets attacker_destroyed=true; defender_crystal_loss=0",
                f"AbilityType.{ab}",
            ],
            [
                f"{name} Bio character; play Release Mutagen → has_mutagen_flag=true.",
                "Opponent attacks with ATK < effective DEF.",
            ],
            ["Resolve defense."],
            [
                "Attacker destroyed; attacker pays crystal cost.",
                f"Defender {name} survives.",
                "When Lab Bloater later destroyed: pay_cost=false (no crystal loss to owner).",
            ],
            [],
        )

    elif ab == "MUTAGEN_ATK_BOOST_VS_AFFINITIES":
        bonus = p.get("bonus", 0)
        affs = p.get("affinities", [])
        add(
            f"{name}: +{bonus} ATK vs {affs} when has_mutagen_flag",
            ["BattleResolver._get_effective_atk() checks has_mutagen_flag", f"AbilityType.{ab}"],
            [f"Apply Release Mutagen to {name}.", f"Target defender affinity in {affs or ['configured affinities']}."],
            ["Attack with mutagen flag."],
            [f"attacker_atk_used == {atk + bonus} when affinity matches and has_mutagen_flag."],
            [],
        )

    elif ab == "MUTAGEN_IMMEDIATE_ATTACK":
        add(
            f"{name}: extra attack capability with mutagen (Lab Crawler)",
            ["TurnManager / GameState mutagen attack grant", f"AbilityType.{ab}"],
            ["Release Mutagen on Lab Crawler; attacks_remaining=2."],
            ["Attack twice in same turn after mutagen."],
            ["Second attack allowed same turn with mutagen flag."],
            [],
        )

    elif ab == "IMMUNE_ZERO_COST_TRAPS":
        add(
            f"{name}: nullifies 0-cost trap activation",
            [
                "BattleResolver._resolve_trap: IMMUNE_ZERO_COST_TRAPS → special_trigger=trap_nullified",
                f"AbilityType.{ab}",
            ],
            ["Opponent Trap Hole (cost 0) face-down.", f"{name} as attacker."],
            ["Attack trap cell."],
            [
                "result.special_trigger == 'trap_nullified'.",
                "Attacker NOT destroyed by trap.",
                "Trap still consumed from grid.",
            ],
            [],
        )

    elif ab == "IMMUNE_TO_TRAPS":
        add(
            f"{name}: immune to all traps",
            ["BattleResolver._resolve_trap IMMUNE_TO_TRAPS", f"AbilityType.{ab}"],
            ["Any trap including Spike Trap (1500 cost).", f"{name} attacks trap."],
            ["Resolve trap trigger."],
            ["special_trigger=='trap_nullified'; attacker survives."],
            [],
        )

    elif ab == "IMMUNE_TO_TECH_CARDS":
        add(
            f"{name}: unaffected by Tech",
            ["GameBoard tech target filters skip IMMUNE cards", f"AbilityType.{ab}"],
            ["Opponent plays Accident targeting {name}.".format(name=name)],
            ["Select {name} if UI allows.".format(name=name)],
            [f"{name} not destroyed; no stat change from Tech."],
            [],
        )

    elif ab == "NEGATE_ZERO_COST_TRAPS_BOTH":
        add(
            f"{name}: passive — all 0-cost traps nullified on both fields",
            ["BattleResolver scans both grids for NEGATE_ZERO_COST_TRAPS_BOTH", f"AbilityType.{ab}"],
            [f"{name} face-up.", "Both players have 0-cost traps."],
            ["Attack any 0-cost trap."],
            ["trap_nullified for both sides while Electrogazer face-up."],
            [],
        )

    elif ab == "REDIRECT_DESTRUCTION_TO_ALLY":
        aff = p.get("affinity", "?")
        add(
            f"{name}: redirect own destruction to other {aff} ally",
            [
                "TurnManager._apply_battle_result prompts own_divine_character_redirect",
                "GameBoard destroys chosen ally instead",
                f"AbilityType.{ab}",
            ],
            [
                f"{name} would be destroyed.",
                f"Another face-up {aff} ally on field (not REDIRECT card).",
            ],
            ["Trigger destruction; choose redirect target."],
            [f"{name} survives; chosen {aff} ally destroyed and owner pays cost."],
            [],
        )

    elif ab == "SACRIFICE_FOR_CARD_TYPE":
        needle = p.get("name_contains", "")
        add(
            f"{name}: sacrifice self to save '{needle}' card",
            ["TurnManager._apply_battle_result SACRIFICE_FOR_CARD_TYPE prompt", f"AbilityType.{ab}"],
            [
                f"Face-up {name} on field.",
                f"Ally defender name contains '{needle}' would be destroyed.",
            ],
            ["Accept sacrifice prompt."],
            [f"{name} destroyed instead; '{needle}' ally survives."],
            [],
        )

    elif ab == "INTERCEPT_ALLY_ATTACK":
        aff = p.get("affinity", "?")
        add(
            f"{name}: swap position with attacked {aff} ally",
            ["TurnManager pre-attack INTERCEPT_ALLY_ATTACK prompt", f"AbilityType.{ab}"],
            [f"Another {aff} ally targeted.", f"{name} on field."],
            ["When ally attacked, accept intercept."],
            ["Positions swap; original target unchanged."],
            [],
        )

    elif ab == "OPTIONAL_CRYSTAL_PAY_ATK_BOOST":
        pc, ba = p.get("cost", 0), p.get("atk", 0)
        add(
            f"{name}: optional pay {pc} crystals for +{ba} ATK",
            ["TurnManager pre-battle crystal prompt OPTIONAL_CRYSTAL_PAY_ATK_BOOST", f"AbilityType.{ab}"],
            [f"Player crystals ≥ {pc} and < {pc} for decline test."],
            ["Accept prompt / Decline prompt in separate runs."],
            [
                f"Accept: attacker_atk_used includes +{ba}; crystals -= {pc}.",
                "Decline: no bonus; no payment.",
            ],
            [],
        )

    elif ab == "PERM_DEF_BOOST_ON_DEFEND":
        b = p.get("def", p.get("bonus", 0))
        add(
            f"{name}: +{b} DEF permanently after successful defend",
            ["BattleResolver._apply_defend_effects", f"AbilityType.{ab}"],
            [f"Opponent ATK < {name} DEF."],
            ["Defend successfully."],
            [f"defender.current_def increased by {b} permanently."],
            [],
        )

    elif ab == "PERM_DEF_BOOST_PER_ATTACK_SURVIVE":
        b = p.get("def", 2)
        add(
            f"{name}: +{b} DEF permanently each attack survived",
            ["TurnManager._apply_post_battle_effects", f"AbilityType.{ab}"],
            [f"{name} attacks and survives (wins or ties without self-destruct)."],
            ["Complete attack."],
            [f"perm_def_bonus += {b} (or current_def += {b})."],
            [],
        )

    elif ab == "PERM_ATK_LOSS_PER_ATTACK":
        amt = p.get("amount", 0)
        add(
            f"{name}: -{amt} ATK permanently after each attack",
            ["TurnManager._apply_post_battle_effects", f"AbilityType.{ab}"],
            [f"{name} attacks."],
            ["Resolve any battle result."],
            [f"current_atk -= {amt} (floored at 0)."],
            [],
        )

    elif ab == "PERM_ATK_LOSS_PER_OWN_TURN":
        amt = p.get("amount", 0)
        add(
            f"{name}: -{amt} ATK at end of own turn (face-up)",
            ["TurnManager._end_turn match PERM_ATK_LOSS_PER_OWN_TURN", f"AbilityType.{ab}"],
            [f"{name} face-up at end of owner's turn."],
            ["End turn."],
            [f"current_atk -= {amt}."],
            [],
        )

    elif ab == "PERM_ATK_BOOST_PER_SURVIVE_OPP_TURN":
        amt = p.get("atk", 0)
        add(
            f"{name}: +{amt} ATK when alive at end of opponent turn",
            ["TurnManager._apply_end_of_turn_boosts", f"AbilityType.{ab}"],
            [f"{name} survives opponent's turn face-up."],
            ["End opponent turn / start own turn."],
            [f"current_atk += {amt}."],
            [],
        )

    elif ab == "ATK_ZERO_AFTER_WIN":
        add(
            f"{name}: ATK→0 permanently after winning battle",
            ["TurnManager._apply_post_battle_effects when defender_destroyed", f"AbilityType.{ab}"],
            [f"{name} wins attack."],
            ["Destroy defender."],
            ["current_atk == 0 after battle."],
            [],
        )

    elif ab == "ONE_USE_EXTRA_ATTACK_ON_KILL":
        add(
            f"{name}: one extra attack after destroying a card",
            ["TurnManager._apply_post_battle_effects; flag 'extra_kill_used'", f"AbilityType.{ab}"],
            [f"{name} destroys defender."],
            ["Check attacks_remaining."],
            ["attacks_remaining += 1 once; flag extra_kill_used set."],
            [],
        )

    elif ab == "EXTRA_ATTACK_VS_REVEALED":
        add(
            f"{name}: extra attack when target was exposed before attack",
            ["TurnManager requires defender_was_exposed==true", f"AbilityType.{ab}"],
            ["Defender face-up before attack selected."],
            [f"Attack with {name}."],
            ["attacks_remaining += 1 once per turn; flag extra_vs_revealed_used."],
            [],
        )

    elif ab == "EXTRA_ATTACK_ON_DEAD_END":
        add(
            f"{name}: extra attack after hitting dead_end (once/turn)",
            ["TurnManager._apply_post_battle_effects", f"AbilityType.{ab}"],
            ["Target cell is dead_end."],
            ["Attack dead_end."],
            ["attacks_remaining += 1; flag extra_deadend_turn."],
            [],
        )

    elif ab == "ONE_USE_EXTRA_ATTACK_ON_DEAD_END":
        add(
            f"{name}: one lifetime extra attack on dead_end",
            ["TurnManager flag extra_deadend_used", f"AbilityType.{ab}"],
            ["First dead_end attack."],
            ["Attack dead_end once."],
            ["One bonus attack total per card lifetime."],
            [],
        )

    elif ab == "MULTI_ATTACK_VS_NON_CHARACTER":
        mx = p.get("max_attacks", 3)
        add(
            f"{name}: up to {mx} attacks when targeting non-characters",
            ["TurnManager._apply_post_battle_effects MULTI_ATTACK_VS_NON_CHARACTER", f"AbilityType.{ab}"],
            ["Targets: dead_end, trap in sequence."],
            [f"Chain attacks on non-character cells."],
            [f"Up to {mx} attacks same turn when targeting non-characters."],
            [],
        )

    elif ab == "LOCK_TARGET_ON_ATTACK":
        add(
            f"{name}: lock attacked character from attacking next turn",
            ["TurnManager sets defender.cannot_attack_until = turn_number+1", f"AbilityType.{ab}"],
            ["Defender survives attack."],
            [f"Attack character with {name}."],
            ["Target cannot_attack_until prevents selection next turn."],
            [],
        )

    elif ab == "LOCK_SELF_AFTER_ATTACK":
        add(
            f"{name}: cannot attack next turn after attacking",
            ["cannot_attack_until = turn_number+2", f"AbilityType.{ab}"],
            [f"{name} attacks."],
            ["End turn; next own turn try to attack with {name}.".format(name=name)],
            [f"{name} not selectable as attacker next turn."],
            [],
        )

    elif ab == "LOCK_ATTACKER_ON_DEFEND":
        add(
            f"{name}: attacker locked after defending",
            ["TurnManager post-defend lock", f"AbilityType.{ab}"],
            [f"Opponent attacks; {name} defends successfully."],
            ["Next opponent turn: same attacker cannot attack."],
            ["Attacker cannot_attack_until set."],
            [],
        )

    elif ab == "LOCK_ATTACKER_ON_DESTROYED":
        add(
            f"{name}: destroyer cannot attack rest of turn",
            ["TurnManager sets attacks_remaining=0", f"AbilityType.{ab}"],
            [f"Opponent destroys {name}."],
            ["Same turn: attacker has attacks_remaining=0."],
            [],
            [],
        )

    elif ab == "REVEAL_ON_ANY_ATTACK":
        add(
            f"{name}: reveal 1 opponent hidden cell after any attack",
            ["TurnManager emit awaiting_target_selection opponent_any_hidden", f"AbilityType.{ab}"],
            [f"{name} attacks any target."],
            ["Complete target selection UI."],
            ["One opponent face-down cell becomes face_up."],
            [],
        )

    elif ab == "REVEAL_ON_WIN":
        add(
            f"{name}: reveal on win only",
            ["TurnManager when defender_destroyed", f"AbilityType.{ab}"],
            ["Win and lose scenarios."],
            ["Attack character / dead_end."],
            ["Reveal prompt only when defender_destroyed."],
            [],
        )

    elif ab == "REVEAL_ON_DEAD_END_ATTACK":
        add(
            f"{name}: reveal after dead_end attack",
            ["TurnManager when defender.card_type=='dead_end'", f"AbilityType.{ab}"],
            ["dead_end target."],
            ["Attack empty cell."],
            ["Reveal selection triggered."],
            [],
        )

    elif ab == "REVEAL_ON_TRAP_ATTACK":
        add(
            f"{name}: reveal after trap attack",
            ["TurnManager when special_trigger trap_*", f"AbilityType.{ab}"],
            ["Opponent trap face-down."],
            ["Attack trap."],
            ["Reveal prompt after trap resolution."],
            [],
        )

    elif ab == "REVEAL_ADJACENT_AFTER_ATTACK":
        add(
            f"{name}: reveal adjacent cell after attack (Scout Probe)",
            ["TurnManager after battle if attacker survived", f"AbilityType.{ab}"],
            [f"{name} survives attack."],
            ["Select adjacent hidden cell."],
            ["Adjacent cell revealed."],
            [],
        )

    elif ab == "CRYSTAL_GAIN_ON_DEAD_END_ATTACK":
        amt = p.get("amount", 0)
        add(
            f"{name}: +{amt} crystals on dead_end attack",
            ["TurnManager GameState.gain_crystals", f"AbilityType.{ab}"],
            [f"Record crystals before attack."],
            ["Attack dead_end."],
            [f"crystals[player] += {amt}."],
            [],
        )

    elif ab == "CRYSTAL_GAIN_ON_OPP_REVEAL":
        amt = p.get("amount", 0)
        add(
            f"{name}: +{amt} crystals when opponent cell revealed",
            ["GameState reveal hooks / CardRuleEngine", f"AbilityType.{ab}"],
            [f"{name} face-up on field."],
            ["Reveal any opponent cell (Spy/Radar/attack reveal)."],
            [f"Owner gains {amt} crystals per reveal event."],
            [],
        )

    elif ab == "CRYSTAL_RECOVER_ON_BIG_LOSS":
        th, amt = p.get("threshold", 500), p.get("amount", 300)
        add(
            f"{name}: recover {amt} if single loss ≥{th}",
            ["GameState crystal loss hook", f"AbilityType.{ab}"],
            [f"{name} on field."],
            [f"Lose ≥{th} crystals in one event (destroy high-cost card)."],
            [f"Net loss reduced by {amt} (gain {amt} back)."],
            [],
        )

    elif ab == "OPPONENT_EXTRA_CRYSTAL_LOSS":
        extra = p.get("amount", 0)
        add(
            f"{name}: opponent loses +{extra} on every crystal loss",
            ["GameState.lose_crystals opponent passives", f"AbilityType.{ab}"],
            [f"{name} face-up on opponent's field... (owner is player with Grave Worm)."],
            ["Opponent loses crystals from any source."],
            [f"Each loss event: additional -{extra} crystals."],
            [],
        )

    elif ab == "DEFEND_DRAIN_ATTACKER":
        amt = p.get("drain_amount", p.get("amount", 0))
        add(
            f"{name}: drain {amt} crystals from attacker on defend",
            ["BattleResolver._apply_defend_effects", f"AbilityType.{ab}"],
            [f"Opponent attacks; {name} wins defense."],
            ["Resolve battle."],
            [f"result.attacker_crystal_loss includes +{amt} (or separate drain message)."],
            [],
        )

    elif ab == "DEFEND_PERM_DEBUFF_ATTACKER_ATK_DEF":
        da = p.get("atk", p.get("amount", 0))
        dd = p.get("def", da)
        add(
            f"{name}: attacker permanently -{da} ATK / -{dd} DEF on defend",
            ["BattleResolver._apply_defend_effects", f"AbilityType.{ab}"],
            [f"{name} defends successfully."],
            ["Resolve defense."],
            [f"attacker.current_atk -= {da}; attacker.current_def -= {dd}."],
            [],
        )

    elif ab == "ONE_USE_PERM_DEBUFF_ATTACKER_ATK":
        amt = p.get("atk", p.get("amount", 0))
        add(
            f"{name}: once, defender permanently -{amt} ATK to attacker",
            ["BattleResolver._apply_defend_effects; one_use_def_boost_used flag", f"AbilityType.{ab}"],
            ["First defend only."],
            ["Defend twice across battles."],
            [f"First: attacker -{amt} ATK permanent; second: no debuff."],
            [],
        )

    elif ab == "ONE_USE_DEFEND_MORPH":
        dl, ag = p.get("def_loss", p.get("def", 40)), p.get("atk_gain", p.get("atk", 40))
        add(
            f"{name}: once after defend: -{dl} DEF, +{ag} ATK permanently",
            ["BattleResolver._apply_defend_effects ONE_USE_DEFEND_MORPH", f"AbilityType.{ab}"],
            [f"{name} defends successfully first time."],
            ["Win defense."],
            [f"current_def -= {dl}; current_atk += {ag}; one_use_def_boost_used=true."],
            [],
        )

    elif ab == "ONE_USE_COPY_STATS_ON_SURVIVE":
        add(
            f"{name}: once, copy battled card stats on mutual survival",
            ["TurnManager when both survive; copy_stats_used flag", f"AbilityType.{ab}"],
            ["Tie or mutual survive scenario."],
            ["Battle where neither destroyed."],
            ["perm_atk_bonus += defender.current_atk; perm_def_bonus += defender.current_def."],
            [],
        )

    elif ab == "ONE_USE_SURVIVE_DESTRUCTION":
        add(
            f"{name}: once per card, survive destruction",
            ["TurnManager/GameState destruction intercept", f"AbilityType.{ab}"],
            [f"{name} would be destroyed first time."],
            ["Trigger destruction."],
            ["Card remains; one-use flag consumed."],
            [],
        )

    elif ab == "TEMP_ATK_BOOST_OWN_TURN_START":
        b = p.get("atk", 0)
        add(
            f"{name}: +{b} temp ATK at start of own turn",
            ["TurnManager turn start clears/applies temp bonuses", f"AbilityType.{ab}"],
            [f"{name} on field."],
            ["Start own turn."],
            [f"temp_atk_bonus == {b} until end of turn."],
            [],
        )

    elif ab == "TEMP_BOOST_ON_OPP_TECH":
        ba, bd = p.get("atk", 5), p.get("def", 5)
        add(
            f"{name}: +{ba}/+{bd} temp when opponent plays Tech",
            ["TurnManager.play_tech_card TEMP_BOOST_ON_OPP_TECH loop", f"AbilityType.{ab}"],
            [f"{name} face-up on opponent of tech player."],
            ["Opponent plays any Tech."],
            [f"temp_atk_bonus += {ba}; temp_def_bonus += {bd} until turn end."],
            [],
        )

    elif ab == "SWAP_ATK_DEF_PER_OPP_TURN":
        add(
            f"{name}: swap ATK/DEF at end of each opponent turn",
            ["TurnManager._apply_end_of_turn_boosts", f"AbilityType.{ab}"],
            [f"{name} face-up."],
            ["End opponent turn."],
            ["current_atk and current_def values swapped."],
            [],
        )

    elif ab == "HALVE_DEF_ON_FIRST_EXPOSE":
        add(
            f"{name}: halve DEF permanently on first reveal",
            ["GameState.reveal_card hook", f"AbilityType.{ab}"],
            [f"{name} face-down."],
            ["Reveal {name}.".format(name=name)],
            [f"current_def = {df // 2} (half of {df}, permanent)."],
            [],
        )

    elif ab == "DESTROY_SELF_AT_END_OF_EXPOSE_TURN":
        add(
            f"{name}: destroyed end of turn when first revealed",
            ["TurnManager._end_turn expose_destroy_pending", f"AbilityType.{ab}"],
            [f"Reveal {name} this turn.".format(name=name)],
            ["End turn without destroying in battle."],
            ["Card destroyed pay_cost=false at turn end."],
            [],
        )

    elif ab == "VENOM_FLAG_END_OF_TURN":
        add(
            f"{name}: end of own turn apply venom to opponent face-up card",
            ["TurnManager._end_turn random venom target", f"AbilityType.{ab}"],
            [f"{name} face-up; opponent has face-up character."],
            ["End own turn."],
            ['Random opponent face-up character gets "venom" in flags array.'],
            [],
        )

    elif ab == "TURN_START_COIN_FLIP_FLAG":
        add(
            f"{name}: turn start coin → venom or mutagen flag on selected opponent card",
            ["TurnManager turn start ability", f"AbilityType.{ab}"],
            [f"{name} on field at turn start.", "Opponent face-up character selected."],
            ["Start turn; select target; resolve coin."],
            ["Heads: venom flag; Tails: mutagen string in flags (NOT has_mutagen_flag)."],
            [],
        )

    elif ab == "SELF_DEBUFF_ON_ATTACK_AND_DEFEND":
        aa, da = p.get("atk", 3), p.get("def", 3)
        add(
            f"{name}: once -{aa} ATK on attack, once -{da} DEF on defend",
            ["TurnManager._apply_post_battle_effects SELF_DEBUFF", f"AbilityType.{ab}"],
            ["First attack and first defend."],
            ["Execute both in separate battles."],
            [f"After first attack: current_atk -= {aa}. After first defend: current_def -= {da}."],
            [],
        )

    elif ab == "OPTIONAL_CRYSTAL_PAY_DESTROY_OPPONENT":
        pc = p.get("cost", 1000)
        add(
            f"{name}: pay {pc} crystals to destroy defender (no opp crystal loss)",
            [
                "TurnManager pre-battle OPTIONAL_CRYSTAL_PAY_DESTROY_OPPONENT",
                "GameState.place_dead_end on accept; attack_aborted signal",
                f"AbilityType.{ab}",
            ],
            [
                f"crystals[player] >= {pc}.",
                "Defender is character.",
            ],
            [
                f"Accept prompt: pay {pc}; defender cell → dead_end.",
                "Opponent does NOT lose crystals.",
                "Skip prompt: normal battle proceeds.",
            ],
            [],
        )

    elif ab == "OPTIONAL_CRYSTAL_PAY_DEF_BOOST":
        pc, bd = p.get("cost", 1000), p.get("def", 60)
        add(
            f"{name}: optional pay {pc} for +{bd} DEF during battle",
            ["TurnManager OPTIONAL_CRYSTAL_PAY_DEF_BOOST on defender side", f"AbilityType.{ab}"],
            [f"Union/card defends; crystals>={pc}."],
            ["Accept/decline DEF boost prompt on battle overlay."],
            [f"Accept: effective DEF +{bd}; crystals -= {pc}."],
            [],
        )

    elif ab == "DOUBLE_STATS_VS_AFFINITY":
        aff = p.get("affinity", "?")
        add(
            f"{name}: double ATK/DEF vs {aff}",
            ["BattleResolver doubles get_effective_atk/def when affinity matches", f"AbilityType.{ab}"],
            [f"Battle vs {aff} opponent."],
            ["Resolve battle."],
            [
                f"attacker_atk_used == base_atk * 2 (when attacking {aff}).",
                f"defender_def_used == base_def * 2 (when defending vs {aff}).",
            ],
            [],
        )

    elif ab == "FIELD_ATK_BOOST_OWN_AFFINITY":
        aff = p.get("affinity", "?")
        ba = p.get("atk", 0)
        add(
            f"{name}: aura +{ba} ATK to own {aff} characters",
            ["BattleResolver field scan FIELD_ATK_BOOST_OWN_AFFINITY", f"AbilityType.{ab}"],
            [f"{name} face-up.", f"Other face-up {aff} ally attacks."],
            ["Ally attacks while aura source exposed."],
            [f"Ally attacker_atk_used includes +{ba} from {name} aura."],
            [],
        )

    elif ab == "GAIN_HALF_STATS_ON_SURVIVE":
        add(
            f"{name}: gain half of opponent ATK/DEF after surviving battle",
            ["TurnManager._apply_post_battle_effects GAIN_HALF_STATS_ON_SURVIVE", f"AbilityType.{ab}"],
            [f"{name} attacks character and survives."],
            ["Complete battle."],
            ["current_atk += defender.current_atk/2; current_def += defender.current_def/2."],
            [],
        )

    elif ab == "MULTI_ATTACK_ANY_WITH_ATK_LOSS":
        mx, loss = p.get("max_attacks", 3), p.get("atk_loss", 5)
        add(
            f"{name}: up to {mx} attacks, -{loss} ATK per attack",
            ["TurnManager MULTI_ATTACK_ANY_WITH_ATK_LOSS", f"AbilityType.{ab}"],
            [f"attacks_remaining sufficient."],
            [f"Chain {mx} attacks same turn."],
            [f"Each attack: current_atk -= {loss}; extra attack while multi_attack_count < {mx}."],
            [],
        )

    elif ab == "STANCE_FIXED_STATS":
        add(
            f"{name}: fixed ATK/DEF by stance",
            ["BattleResolver STANCE_FIXED_STATS uses atk_atk/def_def params", f"AbilityType.{ab}"],
            [f"Attack and defend scenarios separately."],
            ["Resolve each battle."],
            ["Attacking uses atk_atk/def_atk params; defending uses def_def params from ability_params."],
            [],
        )

    elif ab == "HALVE_ATK_ADD_TO_DEF_ON_DEFEND":
        add(
            f"{name}: on defend, halve ATK and add to DEF permanently",
            ["BattleResolver._apply_defend_effects HALVE_ATK_ADD_TO_DEF_ON_DEFEND", f"AbilityType.{ab}"],
            [f"{name} defends and survives."],
            ["Resolve defense."],
            ["current_atk halved; current_def += halved amount."],
            [],
        )

    elif ab == "IMMUNE_IF_OWN_SAME_AFFINITY_FACE_UP":
        aff = p.get("affinity", card["affinity"])
        add(
            f"{name}: cannot be destroyed while another {aff} ally face-up",
            ["BattleResolver post-compare IMMUNE_IF_OWN_SAME_AFFINITY_FACE_UP", f"AbilityType.{ab}"],
            [f"Another face-up {aff} ally on field.", f"{name} would be destroyed."],
            ["Resolve battle that would destroy defender."],
            ["defender_destroyed reverted to false; defender_crystal_loss=0."],
            [],
        )

    elif ab == "DESTROY_SELF_AFTER_BATTLE":
        add(
            f"{name}: self-destroy after battle (no crystal loss)",
            ["TurnManager DESTROY_SELF_AFTER_BATTLE pay_cost=false", f"AbilityType.{ab}"],
            [f"{name} completes any battle surviving."],
            ["Finish battle resolution."],
            ["Attacker destroyed via GameState.destroy_card pay_cost=false."],
            [],
        )

    elif ab == "ATK_PENALTY_VS_DEAD_END":
        pen = p.get("penalty", 50)
        add(
            f"{name}: -{pen} ATK permanently when attacking dead_end",
            ["TurnManager ATK_PENALTY_VS_DEAD_END", f"AbilityType.{ab}"],
            [f"{name} attacks dead_end."],
            ["Complete attack."],
            [f"current_atk -= {pen} permanently."],
            [],
        )

    elif ab == "MULTI_ATTACK_ANY":
        mx = p.get("max_attacks", 2)
        add(
            f"{name}: up to {mx} attacks per turn",
            ["TurnManager MULTI_ATTACK_ANY", f"AbilityType.{ab}"],
            [f"attacks_remaining initial value noted."],
            [f"Perform up to {mx} attacks."],
            ["extra attacks granted while multi_attack_count < max_attacks."],
            [],
        )

    elif ab == "END_OF_TURN_COIN_FLIP_STAT_BOOST":
        ba, bd = p.get("atk", 10), p.get("def", 10)
        add(
            f"{name}: end of turn coin +{ba} ATK or +{bd} DEF permanent",
            ["TurnManager._end_turn END_OF_TURN_COIN_FLIP_STAT_BOOST", f"AbilityType.{ab}"],
            [f"{name} face-up at end of own turn."],
            ["End turn."],
            [f"Heads: current_atk += {ba}; Tails: current_def += {bd}."],
            [],
        )

    elif ab == "CANNOT_ATTACK_IF_NON_AFFINITY_ON_FIELD":
        add(
            f"{name}: blocked from attacking if disallowed affinity on field",
            ["TurnManager attack validation CANNOT_ATTACK_IF_NON_AFFINITY_ON_FIELD", f"AbilityType.{ab}"],
            ["Place disallowed affinity face-up on own field."],
            [f"Try select {name} as attacker."],
            ["Attack blocked or unit not selectable."],
            [],
        )

    elif ab == "COIN_FLIP_NULLIFY_ON_DEFEND":
        add(
            f"{name}: coin flip nullifies attack on heads",
            ["BattleResolver pre-compare COIN_FLIP_NULLIFY_ON_DEFEND", f"AbilityType.{ab}"],
            [f"Opponent attacks {name}."],
            ["Resolve coin."],
            ["Heads: no destruction; Tails: normal battle."],
            [],
        )

    elif ab == "ONE_USE_DESTROY_BY_AFFINITY":
        a1, a2 = p.get("aff1", "?"), p.get("aff2", "?")
        add(
            f"{name}: once destroy defender matching {a1} or {a2} (no crystal loss)",
            ["BattleResolver ONE_USE_DESTROY_BY_AFFINITY", f"AbilityType.{ab}"],
            [f"Defender affinity in ({a1},{a2}); one_use_atk_boost_used=false."],
            ["Attack matching defender."],
            [
                "defender_destroyed=true; defender_crystal_loss=0.",
                "one_use_atk_boost_used=true; second use does not auto-destroy.",
            ],
            [],
        )

    elif ab == "ATK_PENALTY_IF_NO_NAME_ALLY":
        needle = p.get("name_contains", "")
        pen = p.get("penalty", 0)
        add(
            f"{name}: -{pen} ATK if no ally name contains '{needle}'",
            ["BattleResolver._get_effective_atk ATK_PENALTY_IF_NO_NAME_ALLY", f"AbilityType.{ab}"],
            ["No other card with matching name on field."],
            [f"Attack with {name}."],
            [f"attacker_atk_used == {max(0, atk - pen)}."],
            [],
        )

    elif ab == "ATK_DEF_BONUS_IF_OWN_REVEALED_GTE":
        mn, ba, bd = p.get("min_revealed", 15), p.get("atk", 100), p.get("def", 100)
        add(
            f"{name}: +{ba}/+{bd} if ≥{mn} own cells revealed",
            ["BattleResolver ATK_DEF_BONUS_IF_OWN_REVEALED_GTE", f"AbilityType.{ab}"],
            [f"Reveal ≥{mn} cells on own grid (including dead_ends face-up)."],
            ["Battle with card."],
            [f"Stats include +{ba} ATK / +{bd} DEF when threshold met."],
            [],
        )

    elif ab == "PERM_BOOST_END_OF_TURN":
        ba, bd = p.get("atk", 0), p.get("def", 0)
        add(
            f"{name}: +{ba}/+{bd} permanent at end of turn",
            ["TurnManager._apply_end_of_turn_boosts PERM_BOOST_END_OF_TURN", f"AbilityType.{ab}"],
            [f"{name} face-up."],
            ["End own turn."],
            [f"current_atk += {ba}; current_def += {bd}."],
            [],
        )

        add(
            f"{name}: standard ATK vs DEF (no ability)",
            ["BattleResolver standard comparison only", "AbilityType.NONE"],
            [f"{name} ATK={atk} DEF={df}."],
            ["Attack Wandering Swordsman (60/60) or suitable target."],
            [
                "Pure stat comparison; no ability_triggered flags.",
                "Crystal loss equals destroyed card's cost.",
            ],
            [],
        )

    else:
        add(
            f"{name}: ability {ab} functional smoke test",
            [f"CharacterData.AbilityType.{ab}", "See BattleResolver.gd / TurnManager.gd"],
            [f"Card placed; ability_params={p}."],
            ["Trigger battle/turn/tech condition per description."],
            [f"Behavior matches CardDatabase description: {card.get('description', '')}"],
            [],
        )

    return cases


def build_trap_tests(card: dict) -> list[str]:
    name = card["name"]
    sid = slug(name)
    eff = card["effect"]
    p = parse_params(card.get("params_raw", ""))
    cost = card["cost"]
    cases = []
    n = 1

    def add(desc, impl, pre, steps, expected, verify=None):
        nonlocal n
        cid = f"TC-FUNC-{sid}-{n:03d}"
        n += 1
        cases.append(fmt_case(cid, desc, impl, BASE_PRE + pre, steps, expected, verify))

    TRAP_IMPL = {
        "DRAIN_ATTACKER_CRYSTALS": (
            f"Attacker loses {p.get('amount', 800)} crystals",
            [
                "TurnManager._handle_trap_effect DRAIN_ATTACKER_CRYSTALS",
                f"TrapEffectType.{eff}",
                f"effect_params: {p}",
            ],
            [
                f"Trap {name} face-down cost={cost}.",
                f"Attacker player crystals tracked (start 5000).",
            ],
            ["Attack trap cell."],
            [
                f"Attacker crystals -= {p.get('amount', 800)} via GameState.lose_crystals.",
                f"Trap owner pays trap cost {cost} (defender_crystal_loss in resolver).",
                "Trap cell becomes dead_end.",
            ],
        ),
        "NULLIFY_ATTACK_ATK_DEBUFF": (
            "Foul Gas: -5 ATK temp to all attacker face-up chars",
            ["TurnManager TEMP_DEBUFF_ALL_ATTACKER_CHARS", f"TrapEffectType.{eff}"],
            ["Attacker has multiple face-up characters.", f"{name} on field."],
            ["Attack trap."],
            [f"Each attacker-side face-up character temp_atk_bonus -= {p.get('amount', 5)} until turn end."],
        ),
        "NULLIFY_ATTACK_REVEAL_ADJACENT": (
            "Hostage: reveal adjacent + lock from targeting",
            ["TurnManager NULLIFY_ATTACK_REVEAL_ADJACENT", f"TrapEffectType.{eff}"],
            ["Hidden cells adjacent to trap.", f"{name} at center of cluster."],
            ["Attack trap."],
            [
                "All adjacent opponent cells revealed.",
                "Positions added to GameState.locked_attack_positions until turn end.",
            ],
        ),
        "REVEAL_DEFENDING_CHOICE": (
            "Bait: defender chooses own cell to reveal",
            ["TurnManager REVEAL_DEFENDING_CHOICE", f"TrapEffectType.{eff}"],
            [f"{name} face-down."],
            ["Attack trap; defender selects reveal target."],
            ["Chosen cell face_up=true."],
        ),
        "ATTACKER_DISCARD_OR_END_TURN": (
            "Blackmail: discard Tech or end turn",
            ["TurnManager ATTACKER_DISCARD_OR_END_TURN", f"TrapEffectType.{eff}"],
            ["Attacker has Tech in hand.", f"{name} face-down."],
            ["Attack trap; choose branch."],
            [
                "Choice 0: pop Tech from hand OR attacks_remaining=0 if empty.",
                "Choice 1: attacks_remaining=0.",
            ],
        ),
        "FORCE_FRIENDLY_FIRE": (
            "Brainwash: attacker must hit own ally",
            ["TurnManager FORCE_FRIENDLY_FIRE", f"TrapEffectType.{eff}"],
            ["Attacker has own character on field.", f"{name} cost {cost}."],
            ["Attack trap; select own ally as redirected target."],
            ["Second battle resolves against own card."],
        ),
        "DESTROY_ATTACKER": (
            f"Destroy attacker (Spike/Flame Trap)",
            ["TurnManager DESTROY_ATTACKER", f"TrapEffectType.{eff}"],
            [f"{name} face-down.", "Attacker on field."],
            ["Attack trap."],
            [
                "Attacker cell → dead_end.",
                f"Attacker owner pays attacker.crystal_cost.",
            ],
        ),
        "DESTROY_ATTACKER_DEFENDER_PAYS": (
            f"Explosive Barrels: destroy attacker + defender pays {p.get('amount', 'attacker cost')}",
            ["TurnManager DESTROY_ATTACKER_DEFENDER_PAYS", f"TrapEffectType.{eff}"],
            [f"{name} face-down."],
            ["Attack with costly character."],
            [
                "Attacker destroyed and pays cost.",
                f"Defender (trap owner) loses {p.get('amount', 'attacker.crystal_cost')} crystals.",
            ],
        ),
        "HYPNOTIZE_ATTACKER": (
            "Hypnosis: attacker cannot attack next turn",
            ["TurnManager sets cannot_attack_until=turn_number+1", f"TrapEffectType.{eff}"],
            [f"{name} triggered."],
            ["Attack trap."],
            ["attacker.cannot_attack_until prevents next-turn attack."],
        ),
        "LOCK_ATTACKER_REMAINING_ATTACKS": (
            "Echo Barrier: no more attacks this turn",
            ["TurnManager attacks_remaining=0", f"TrapEffectType.{eff}"],
            ["attacks_remaining was 2.", f"{name} on field."],
            ["Attack trap mid-turn."],
            ["attacks_remaining == 0 immediately."],
        ),
        "NULLIFY_ATTACKER_EFFECT": (
            "Snare Trap: attacker ability nullified",
            ["TurnManager effect_nullified_until=turn_number+1", f"TrapEffectType.{eff}"],
            ["Attacker has active ability.", f"{name} on field."],
            ["Attack trap then attack with same unit."],
            ["Ability bonuses suppressed until end of next turn."],
        ),
        "PERMANENT_ATK_DEBUFF": (
            f"Flame Trap: permanent -{p.get('amount', 10)} ATK",
            ["TurnManager PERMANENT_ATK_DEBUFF", f"TrapEffectType.{eff}"],
            [f"{name} triggered."],
            ["Attack trap."],
            [f"attacker.current_atk -= {p.get('amount', 10)} permanently."],
        ),
        "SWAP_ATTACKER_ATK_DEF_TEMP": (
            "Cursed Reflection: swap attacker ATK/DEF until turn end",
            ["TurnManager SWAP_ATTACKER_ATK_DEF_TEMP temp bonuses", f"TrapEffectType.{eff}"],
            ["Attacker with ATK != DEF.", f"{name} on field."],
            ["Attack trap."],
            ["Effective ATK/DEF swapped via temp_atk_bonus/temp_def_bonus."],
        ),
        "NULLIFY_BLOCK_ADJACENT": (
            "Bunker: adjacent cells not targetable",
            ["TurnManager NULLIFY_BLOCK_ADJACENT", f"TrapEffectType.{eff}"],
            [f"{name} on field."],
            ["Attack trap; try attack adjacent cell same turn."],
            ["Adjacent positions in locked_attack_positions."],
        ),
        "FIELD_BOOST_AFFINITY_DEF": (
            f"Alarm: +{p.get('def', 5)} temp DEF to Anima face-up",
            ["TurnManager FIELD_BOOST_AFFINITY_DEF", f"TrapEffectType.{eff}"],
            ["Face-up Anima characters on trap owner's field."],
            ["Trigger trap."],
            [f"Each matching character temp_def_bonus += {p.get('def', 5)}."],
        ),
        "SWAP_ARMORED_NATURE": (
            "Defensive Pheromone: swap with Armored Nature",
            ["TurnManager awaiting own_armored_nature", f"TrapEffectType.{eff}"],
            ["Armored Nature card on field.", f"{name} on field."],
            ["Trigger; select Armored Nature."],
            ["Grid positions of trap and selected card swap."],
        ),
        "CANCEL_ATTACKER_ATTACK": (
            "Decoy Puppet: cancel this attack",
            ["TurnManager CANCEL_ATTACKER_ATTACK", f"TrapEffectType.{eff}"],
            [f"{name} set to trigger on battle."],
            ["Attack while Decoy condition met."],
            ["No battle damage; attack cancelled."],
        ),
        "TEMP_DEF_BOOST_ONE_OWN": (
            f"Hard Scale: +{p.get('def', 5)} temp DEF to chosen own card",
            ["TurnManager TEMP_DEF_BOOST_ONE_OWN", f"TrapEffectType.{eff}"],
            [f"{name} triggered on battle."],
            ["Select own character."],
            [f"Selected card temp_def_bonus += {p.get('def', 5)}."],
        ),
        "COIN_FLIP_2_ATK_DEBUFF": (
            f"Pepper Spray: 2 coins both heads → -{p.get('amount', 5)} ATK",
            ["TurnManager COIN_FLIP_2_ATK_DEBUFF", f"TrapEffectType.{eff}"],
            [f"{name} on field."],
            ["Attack trap; log coin results."],
            [f"Both heads: current_atk -= {p.get('amount', 5)}; else no effect."],
        ),
        "COIN_FLIP_2_LOCK_ATTACKER": (
            "Red Card: 2 heads → cannot attack next turn",
            ["TurnManager COIN_FLIP_2_LOCK_ATTACKER", f"TrapEffectType.{eff}"],
            [f"{name} on field."],
            ["Attack trap."],
            ["Both heads: cannot_attack_until set."],
        ),
        "SELF_DESTROY_TEMP_ATK_BOOST": (
            f"Self-destruct: +{p.get('atk', 10)} ATK then destroy next turn end",
            ["TurnManager SELF_DESTROY_TEMP_ATK_BOOST", f"TrapEffectType.{eff}"],
            [f"{name} on field."],
            ["Choose own character."],
            [f"+{p.get('atk', 10)} temp ATK; destroyed end of next turn pay_cost=false."],
        ),
        "REVEAL_OWN_GAIN_CRYSTAL": (
            f"Street Joke: reveal own hidden +{p.get('amount', 100)} crystals",
            ["TurnManager REVEAL_OWN_GAIN_CRYSTAL", f"TrapEffectType.{eff}"],
            ["Trap owner has face-down non-dead_end cell."],
            ["Attack trap."],
            [f"Random own cell revealed; gain {p.get('amount', 100)} crystals."],
        ),
        "COPY_ATTACKER_EFFECT": (
            "Cursed Reflection copy (legacy enum)",
            ["TurnManager COPY_ATTACKER_EFFECT", f"TrapEffectType.{eff}"],
            [f"{name} — verify actual card mapping in CardDatabase."],
            ["Attack trap."],
            ["Follow TurnManager branch for copy effect."],
        ),
    }

    if eff in TRAP_IMPL:
        desc, impl, pre, steps, expected = TRAP_IMPL[eff]
        add(desc, impl, pre, steps, expected)
    else:
        add(
            f"{name}: trap {eff}",
            [f"TrapEffectType.{eff}", "TurnManager._handle_trap_effect"],
            [f"{name} face-down."],
            ["Attack trap."],
            [card.get("description", "")],
        )

    # Immunity cross-test for 0-cost traps
    if cost == 0:
        add(
            f"{name}: nullified by Huntress/Laser Walker/Electrogazer",
            [
                "BattleResolver._resolve_trap immunity + NEGATE_ZERO_COST_TRAPS_BOTH",
                f"Zero-cost trap cost={cost}",
            ],
            [
                "Case A: Huntress of Green Glade attacks trap.",
                "Case B: Electrogazer face-up on either field.",
            ],
            ["Attack trap under each immunity scenario."],
            [
                "special_trigger=='trap_nullified'.",
                "Attacker not destroyed; crystal drain from DRAIN may still be skipped.",
            ],
        )

    return cases


def build_tech_tests(card: dict) -> list[str]:
    name = card["name"]
    sid = slug(name)
    eff = card["effect"]
    p = parse_params(card.get("params_raw", ""))
    cost = card["cost"]
    cases = []
    n = 1

    def add(desc, impl, pre, steps, expected, verify=None):
        nonlocal n
        cid = f"TC-FUNC-{sid}-{n:03d}"
        n += 1
        cases.append(fmt_case(cid, desc, impl, BASE_PRE + pre, steps, expected, verify))

    TECH_SPECS = {
        "REVEAL_OPPONENT_SQUARE": (
            f"Reveal {p.get('count', 1)} opponent square(s)",
            ["TurnManager.play_tech_card → awaiting_target_selection", f"TechEffectType.{eff}"],
            [f"{name} in hand; phase MODE_SELECT; crystals>={cost}."],
            [f"Play {name}; select {p.get('count', 1)} hidden opponent cell(s)."],
            [
                f"crystals[player] -= {cost}.",
                "Selected cells face_up=true.",
                f"Tech removed from hand; tech_cards_played_this_game updated.",
            ],
        ),
        "OPPONENT_REVEALS_SQUARE": (
            "Tease: opponent reveals 1 own square",
            ["TurnManager OPPONENT_REVEALS_SQUARE", f"TechEffectType.{eff}"],
            [f"{name} in hand."],
            [f"Play {name}."],
            ["Opponent chooses own cell; it reveals."],
        ),
        "OPPONENT_REVEALS_OR_GAINS": (
            f"Bribe: opponent reveal creature for {p.get('crystal_reward', 700)} or pass",
            ["TurnManager OPPONENT_REVEALS_OR_GAINS", f"TechEffectType.{eff}"],
            [f"{name} in hand."],
            [f"Play {name}; opponent accepts or declines."],
            [
                f"Accept: reveal creature + gain {p.get('crystal_reward', 700)} crystals.",
                "Decline: no effect.",
            ],
        ),
        "ADD_MUTAGEN_FLAG": (
            "Release Mutagen: has_mutagen_flag on Bio character",
            [
                "TurnManager.play_tech_card ADD_MUTAGEN_FLAG",
                "GameBoard pending_tech_filter own_bio_character sets has_mutagen_flag=true",
                f"TechEffectType.{eff}",
            ],
            ["Face-down Lab Zombie on own field.", f"{name} in hand."],
            [f"Play {name}; select Bio character."],
            [
                "Card revealed if face-down.",
                "has_mutagen_flag == true (NOT same as flags array 'mutagen').",
                "Enables MUTAGEN_* battle abilities.",
            ],
        ),
        "DIVINE_PROTECTION": (
            "Prayer: Divine characters survive destruction once",
            [
                "TurnManager sets GameState.divine_protection_active[player]=true",
                "TurnManager._apply_battle_result cancels Divine destruction once",
                f"TechEffectType.{eff}",
            ],
            ["Own Divine character would be destroyed.", f"{name} played."],
            [f"Play {name}; trigger destruction on Divine."],
            [
                "First destruction prevented; divine_protection_active cleared.",
                "Second destruction not prevented.",
            ],
        ),
        "REVEAL_ALL_OWN_CHARACTERS": (
            "Great Diplomacy: all own characters face-up",
            ["TurnManager._reveal_all_own", f"TechEffectType.{eff}"],
            ["Multiple own face-down characters.", f"crystals>={cost}."],
            [f"Play {name}."],
            ["All own character cells face_up=true."],
        ),
        "OPPONENT_NEXT_DEFENDER_DESTROYED": (
            "Siege Cannon: next defender auto-destroyed after battle",
            [
                "TurnManager sets siege_cannon_active[player]=true",
                "After battle if defender survives → GameState.destroy_card",
                f"TechEffectType.{eff}",
            ],
            [f"{name} played.", "Opponent defender that would survive battle."],
            [f"Play {name}; attack character where ATK<DEF."],
            [
                "Defender destroyed anyway after battle resolution.",
                "siege_cannon_active cleared after trigger.",
            ],
        ),
        "DESTROY_FACEUP_NO_CRYSTAL_LOSS": (
            "Accident: destroy face-up card without crystal cost",
            ["TurnManager/GameBoard destroy pay_cost=false", f"TechEffectType.{eff}"],
            ["Opponent face-up character.", f"crystals>={cost}."],
            [f"Play {name}; select face-up target."],
            ["Target destroyed; owner crystals unchanged."],
        ),
        "REVIVE_CHARACTER_NO_ATK": (
            "Resurrection: revive character ATK=0 ability NONE",
            ["TurnManager REVIVE_CHARACTER_NO_ATK", f"TechEffectType.{eff}"],
            ["Previously destroyed character tracked.", "Empty cell.", f"crystals>={cost}."],
            [f"Play {name}; place on empty cell."],
            [
                "Card face_up at chosen cell.",
                "current_atk=0; ability_type=NONE.",
                "Second play blocked if once-only enforced.",
            ],
        ),
        "NOT_IMPLEMENTED": (
            f"{name}: NOT_IMPLEMENTED in engine",
            ["TurnManager.play_tech_card → NOT_IMPLEMENTED branch", f"TechEffectType.{eff}"],
            [f"{name} in hand."],
            [f"Play {name}."],
            [
                "GameState.show_center_message('Ability not implemented: ...').",
                f"Designed effect (from CardDatabase): {card.get('description', '')}",
                "No stat change until implemented.",
            ],
        ),
    }

    if eff in TECH_SPECS:
        desc, impl, pre, steps, expected = TECH_SPECS[eff]
        add(desc, impl, pre, steps, expected)
    else:
        add(
            f"{name}: tech {eff}",
            [f"TechEffectType.{eff}", "TurnManager.play_tech_card"],
            [f"{name} in hand during MODE_SELECT."],
            [f"Play {name}."],
            [card.get("description", "")],
        )

    if cost == 0:
        add(
            f"{name}: playable at 0 crystals",
            ["TurnManager.play_tech_card cost check"],
            ["Set crystals[player]=0."],
            [f"Play {name}."],
            ["Tech resolves; crystals remain 0."],
        )

    return cases


def build_union_tests(card: dict) -> list[str]:
    sid = slug(card["name"])
    cases: list[str] = []

    if card.get("ability") == "EXCEL_ONLY_NOT_IN_UNION_DATABASE":
        cases.append(
            fmt_case(
                f"TC-FUNC-{sid}-001",
                f"{card['name']}: pending UnionDatabase implementation",
                [
                    "UnionDatabase.gd — NOT YET REGISTERED",
                    "Excel Demo=Yes with planned ability text",
                ],
                BASE_PRE
                + [
                    f"Card listed in card_data_demo.xlsx Union sheet.",
                    f"Planned formula: {card.get('formula', '?')}",
                ],
                [
                    "Verify card appears in design spreadsheet.",
                    "When added to UnionDatabase.gd, replace with code-derived functional tests.",
                ],
                [
                    f"Excel ability: {card.get('description', '')}",
                    "Engine test blocked until _add() entry exists.",
                ],
            )
        )
        return cases

    card_copy = {**card, "card_type": "Union", "cost": card.get("summon_cost", 0)}
    cases = build_character_tests(card_copy)
    summon = card["summon_cost"]

    header_extra = fmt_case(
        f"TC-FUNC-{sid}-000",
        f"Union summon: {card['name']}",
        [
            "GameBoard._perform_pending_union",
            "UnionDatabase.find_available_unions validation",
            "Once per duel per player",
        ],
        [
            f"summon_cost={summon} crystals available.",
            "Material cells satisfy UnionDatabase material_conditions.",
        ],
        [
            "Enter union mode; select valid materials.",
            f"Pay {summon} crystals; place at anchor cell.",
        ],
        [
            "is_union=true; face_up=true at anchor.",
            "Materials removed pay_cost=false.",
            "_union_summoned_this_duel[player]=true blocks second union.",
        ],
    )
    return [header_extra] + cases


def write_implementation_index(demo_cards: list[dict]) -> None:
    abilities = sorted({c["ability"] for c in demo_cards if c.get("ability")})
    lines = [
        "# Implementation Index (Demo Cards)",
        "",
        "Code paths for functional verification.",
        "",
        "| AbilityType / EffectType | Primary handler |",
        "|------------------------|-----------------|",
        "| ATK/DEF modifiers | `BattleResolver._get_effective_atk/def()` |",
        "| Pre-compare destroy | `BattleResolver._resolve_character_vs_character()` |",
        "| Trap trigger | `BattleResolver._resolve_trap()` → `TurnManager._handle_trap_effect()` |",
        "| Post-attack | `TurnManager._apply_post_battle_effects()` |",
        "| Post-defend | `BattleResolver._apply_defend_effects()` |",
        "| Turn end | `TurnManager._end_turn()` / `_apply_end_of_turn_boosts()` |",
        "| Tech play | `TurnManager.play_tech_card()` → `GameBoard` target filters |",
        "| Union summon | `GameBoard._perform_pending_union()` |",
        "| Field passives | `BattleResolver.calculate_field_bonuses()` |",
        "",
        "## Demo ability types in use",
        "",
    ]
    for ab in abilities:
        lines.append(f"- `{ab}`")
    (OUT / "implementation_index.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    demo_names = load_demo_names()
    excel_unions = load_union_demo_from_excel()
    all_cards = parse_card_database()
    unions_db = parse_union_database()

    by_type: dict[str, list[dict]] = {"Character": [], "Trap": [], "Tech": [], "Union": []}
    for name, card in all_cards.items():
        if name in demo_names:
            by_type[card["card_type"]].append(card)
    by_type["Union"] = merge_union_cards(unions_db, excel_unions)

    for t in by_type:
        by_type[t].sort(key=lambda c: (c.get("ability", c.get("effect", "")), c["name"]))

    builders = {
        "Character": build_character_tests,
        "Trap": build_trap_tests,
        "Tech": build_tech_tests,
        "Union": build_union_tests,
    }

    manifest: list[str] = []
    total = 0

    for card_type, cards in by_type.items():
        parts = [
            f"# {card_type} — Functional Test Cases (Demo = Yes)",
            "",
            "Derived from Godot `CardDatabase.gd` / `UnionDatabase.gd` implementation.",
            "Each case references the exact handler function and enum type.",
            "",
            f"**Cards:** {len(cards)}",
            "",
            "---",
            "",
        ]
        for card in cards:
            parts.append(card_header(card))
            parts.append("")
            tests = builders[card_type](card)
            for tc in tests:
                parts.append(tc)
                parts.append("")
                m = re.search(r"Test Case ID: (TC-FUNC-[^\n]+)", tc)
                if m:
                    manifest.append(m.group(1))
                    total += 1
            parts.append("---")
            parts.append("")

        out_file = OUT / f"{card_type.lower()}_functional_tests.md"
        out_file.write_text("\n".join(parts), encoding="utf-8")
        print(f"Wrote {out_file} ({len(cards)} cards)")

    write_implementation_index(by_type["Character"] + by_type["Union"])
    (OUT / "functional_test_manifest.txt").write_text("\n".join(sorted(manifest)) + "\n", encoding="utf-8")
    print(f"Total functional test cases: {total}")


if __name__ == "__main__":
    main()
